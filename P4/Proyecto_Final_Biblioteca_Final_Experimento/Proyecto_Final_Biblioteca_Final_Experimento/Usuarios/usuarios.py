import re
import Usuarios.crud
import funciones 
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from pathlib import Path

def agregarUsuarios(conexionBD):
    print("\033c")
    texto = f"\n\t\t{funciones.AZUL}...:::: AGREGAR USUARIOS ::::...\n"
    funciones.escritura_lenta_print(texto)

    patron_matricula = r'^\d{1,11}$' 
    patron_nombre = r'^[A-ZÁÉÍÓÚÑa-záéíóúñ\s]{3,50}$'
    patron_correo = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
    patron_carrera = r'^[A-ZÁÉÍÓÚÑa-záéíóúñ\s]{3,50}$'
    patron_cuatrimestre = r'^(?:[1-9]|1[0-1])$'
    modalidades_permitidas = ["CLASICA", "BILINGÜE", "BILINGUE"]

    texto = f"{funciones.AMARILLO}Introducir la matricula del usuario (solo numeros): {funciones.RESET}\033[s"
    funciones.escritura_lenta_print(texto)
    matricula = input("\033[u").strip()
    
    resp = Usuarios.crud.existe_matricula(matricula, conexionBD)
    
    while matricula == "" or not re.match(patron_matricula, matricula) or resp:
        if matricula == "":
            print(f"\n{funciones.ROJO}La matricula no puede estar vacia.{funciones.RESET}")
        elif not re.match(patron_matricula, matricula):
            print(f"\n{funciones.ROJO}Formato invalido. Debe contener solo numeros con una longitud maxima de 11 digitos.{funciones.RESET}")
        elif resp:
            print(f"\n{funciones.ROJO}¡La matricula: {matricula} ya existe! Intente con otra.{funciones.RESET}")
            
        funciones.espereTecla()
        print("\033[u\033[J", end="", flush=True)
        matricula = input().strip()
        resp = Usuarios.crud.existe_matricula(matricula, conexionBD)

    texto = f"{funciones.AMARILLO}Introducir el nombre del usuario: {funciones.RESET}\033[s"
    funciones.escritura_lenta_print(texto)
    name = input("\033[u").strip()
    
    while name == "" or not re.match(patron_nombre, name):
        if name == "":
            print(f"\n{funciones.ROJO}El nombre no puede estar vacio.{funciones.RESET}")
        else:
            print(f"\n{funciones.ROJO}El nombre solo debe contener letras y espacios con una longitud entre 3 y 50 caracteres.{funciones.RESET}")
            
        funciones.espereTecla()
        print("\033[u\033[J", end="", flush=True)
        name = input().strip()
        
    name = name.upper()

    texto = f"{funciones.AMARILLO}Introducir el correo electronico del usuario: {funciones.RESET}\033[s"
    funciones.escritura_lenta_print(texto)
    correo = input("\033[u").strip()

    while correo == "" or not re.match(patron_correo, correo):
        if correo == "":
            print(f"\n{funciones.ROJO}El correo no puede estar vacio.{funciones.RESET}")
        else:
            print(f"\n{funciones.ROJO}Formato de correo invalido, {funciones.VERDE}Ej: usuario@dominio.com{funciones.RESET}")
            
        funciones.espereTecla()
        print("\033[u\033[J", end="", flush=True)
        correo = input().strip()

    correo = correo.lower()

    texto = f"{funciones.AMARILLO}Introducir la carrera del usuario: {funciones.RESET}\033[s"
    funciones.escritura_lenta_print(texto)
    carrera = input("\033[u").strip()

    while carrera == "" or not re.match(patron_carrera, carrera):
        if carrera == "":
            print(f"\n{funciones.ROJO}La carrera no puede estar vacia.{funciones.RESET}")
        else:
            print(f"\n{funciones.ROJO}La carrera solo debe contener letras y espacios con una longitud entre 3 y 50 caracteres.{funciones.RESET}")
            
        funciones.espereTecla()
        print("\033[u\033[J", end="", flush=True)
        carrera = input().strip()

    carrera = carrera.upper()

    texto = f"{funciones.AMARILLO}Introducir el cuatrimestre del usuario: {funciones.RESET}\033[s"
    funciones.escritura_lenta_print(texto)
    cuatrimestre = input("\033[u").strip()

    while cuatrimestre == "" or not re.match(patron_cuatrimestre, cuatrimestre):
        if cuatrimestre == "":
            print(f"\n{funciones.ROJO}El cuatrimestre no puede estar vacio.{funciones.RESET}")
        else:
            print(f"\n{funciones.ROJO}Cuatrimestre invalido. Debe ingresar un numero entero del 1 al 11.{funciones.RESET}")
            
        funciones.espereTecla()
        print("\033[u\033[J", end="", flush=True)
        cuatrimestre = input().strip()

    cuatrimestre = int(cuatrimestre)

    texto = f"{funciones.AMARILLO}Introducir la modalidad del usuario (CLASICA o BILINGÜE): {funciones.RESET}\033[s"
    funciones.escritura_lenta_print(texto)
    modalidad = input("\033[u").strip().upper()

    while modalidad == "" or modalidad not in modalidades_permitidas:
        if modalidad == "":
            print(f"\n{funciones.ROJO}La modalidad no puede estar vacia.{funciones.RESET}")
        else:
            print(f"\n{funciones.ROJO}Modalidad invalida. Debe escribir unicamente CLASICA o BILINGÜE.{funciones.RESET}")
            
        funciones.espereTecla()
        print("\033[u\033[J", end="", flush=True)
        modalidad = input().strip().upper()

    if modalidad == "BILINGUE":
        modalidad = "BILINGÜE"

    respuesta = Usuarios.crud.insertar(matricula, name, correo, carrera, cuatrimestre, modalidad, conexionBD)
    if respuesta: 
        funciones.accionExitosa()
        funciones.menuUsuarios()
    else:
        funciones.accionNoExitosa()
        funciones.menuUsuarios()

def mostrarUsuarios(conexionBD): 
    print("\033c")
    texto = f"{funciones.AZUL}\n\t\t\t\t...:::: MOSTRAR USUARIOS ::::...\n{funciones.AMARILLO}"
    funciones.escritura_lenta_print(texto)
    
    usuarios = Usuarios.crud.consultar(conexionBD)
    
    if len(usuarios) > 0:
        print(f"\t{funciones.AMARILLO}{'Matrícula':<15}{'Nombre':<25}{'Correo':<20}{'Carrera':<25}{'Cuatrimestre':<15}{'Modalidad':<15}{funciones.RESET}\n")
        
        for i in usuarios:
            mat = str(i[0])[:14]
            nombre = str(i[1])[:20]
            correo = str(i[2])[:20]
            carrera = str(i[3])[:20]
            cuatri = str(i[4])[:5]
            modalidad = str(i[5])[:15]
            
            print(f"\t{mat:<15}{nombre:<25}{correo:<20}{carrera:<25}{cuatri:<15}{modalidad:<15}")
            
        print("\n")
        funciones.espereTecla()
        funciones.menuUsuarios()
    else:        
        texto = f"\n{funciones.ROJO}...¡No hay usuarios que mostrar!...{funciones.RESET}\n"
        funciones.escritura_lenta_print(texto)
        funciones.espereTecla()
        funciones.menuUsuarios()

def buscarUsuarios(conexionBD):
    print("\033c")
    texto = f"{funciones.AZUL}\n\t\t\t\t...:::: BUSCAR USUARIOS ::::...\n"
    funciones.escritura_lenta_print(texto)
    
    texto = f"{funciones.AMARILLO}Escribir la matrícula o nombre del usuario: {funciones.RESET}\033[s"
    funciones.escritura_lenta_print(texto)
    criterio = input("\033[u").strip()

    while criterio == "":
        print(f"\n{funciones.ROJO}El nombre o matrícula de búsqueda no puede estar vacío.{funciones.RESET}")
        funciones.espereTecla()
        print("\033[u\033[J", end="", flush=True)
        criterio = input().strip()

    criterio = criterio.upper()
    usuarios = Usuarios.crud.buscar(criterio, conexionBD)
    
    if len(usuarios) > 0:
        print(f"\t{funciones.VERDE}Se encontraron {len(usuarios)} registro(s){funciones.RESET}\n")

        print(f"\t{funciones.AMARILLO}{'Matrícula':<15}{'Nombre':<25}{'Correo':<20}{'Carrera':<25}{'Cuatrimestre':<15}{'Modalidad':<15}{funciones.RESET}\n")
                
        for i in usuarios:
            mat = str(i[0])[:14]
            nombre = str(i[1])[:20]
            correo = str(i[2])[:20]
            carrera = str(i[3])[:20]
            cuatri = str(i[4])[:5]
            modalidad = str(i[5])[:15]
                    
        print(f"\t{mat:<15}{nombre:<25}{correo:<20}{carrera:<25}{cuatri:<15}{modalidad:<15}")
            
        print("\n")
        funciones.espereTecla()
        funciones.menuUsuarios()
    else:        
        texto = f"\n{funciones.ROJO}...¡No se encontró ningún usuario con esa matrícula o nombre!...{funciones.RESET}\n"
        funciones.escritura_lenta_print(texto)
        funciones.espereTecla()
        funciones.menuUsuarios()

def borrarUsuarios(conexionBD):
    print("\033c")
    texto = f"{funciones.AZUL}\n\t\t\t\t...:::: BORRAR USUARIO ::::...\n{funciones.RESET}"
    funciones.escritura_lenta_print(texto)

    patron_matricula = r'^\d{1,11}$' 

    texto = f"{funciones.AMARILLO}Escribir la matricula del usuario a borrar: {funciones.RESET}\033[s"
    funciones.escritura_lenta_print(texto)
    
    matricula = input("\033[u").strip()

    while matricula == "" or not re.match(patron_matricula, matricula):
        print(f"\n{funciones.ROJO}Entrada invalida. Debe ingresar unicamente los numeros de la matricula.{funciones.RESET}")
        funciones.espereTecla()
        print("\033[u\033[J", end="", flush=True)
        matricula = input().strip()

    usuarios = Usuarios.crud.buscar(matricula, conexionBD)
    
    if len(usuarios) > 0:
        cursor = conexionBD.cursor()
        cursor.execute("SELECT COUNT(*) FROM prestamos WHERE id_usuario = %s", (matricula,))
        prestamos_asociados = cursor.fetchone()[0]

        if prestamos_asociados > 0:
            input(f"\n{funciones.ROJO}...¡No se puede borrar! El usuario tiene prestamos registrados en el sistema. Presione Enter...{funciones.RESET}")
        else:
            print(f"\t{funciones.AMARILLO}{'Matricula':<15}{'Nombre':<25}{'Correo':<20}{'Carrera':<25}{'Cuatrimestre':<15}{'Modalidad':<15}{funciones.RESET}\n")
                    
            for i in usuarios:
                mat = str(i[0])[:14]
                nombre = str(i[1])[:20]
                correo = str(i[2])[:20]
                carrera = str(i[3])[:20]
                cuatri = str(i[4])[:5]
                modalidad = str(i[5])[:15]
                        
                print(f"\t{mat:<15}{nombre:<25}{correo:<20}{carrera:<25}{cuatri:<15}{modalidad:<15}")
            
            texto = f"\n{funciones.ROJO}¿Deseas borrar al usuario mostrado (Si/No)? {funciones.RESET}\033[s"
            funciones.escritura_lenta_print(texto)
            opc = input("\033[u").lower().strip()

            while opc not in ["si", "no"]:
                print(f"\n{funciones.ROJO}Opcion invalida. Escriba unicamente SI o NO.{funciones.RESET}")
                funciones.espereTecla()
                print("\033[u\033[J", end="", flush=True)
                opc = input().lower().strip()
                    
            if opc == "si":       
                respuesta = Usuarios.crud.borrar(matricula, conexionBD)
                
                if respuesta:
                    funciones.accionExitosa()
                else:
                    funciones.accionNoExitosa()
    else:        
        input(f"\n{funciones.ROJO}...¡No se encontro ningun usuario registrado con la matricula: {matricula}! Presione Enter...{funciones.RESET}")

    funciones.menuUsuarios()

def limpiarUsuarios(conexionBD):
    print("\033c")
    texto = f"{funciones.AZUL}\n\t\t\t\t...:::: LIMPIAR USUARIOS ::::...\n{funciones.RESET}"
    funciones.escritura_lenta_print(texto)
    
    usuarios = Usuarios.crud.consultar(conexionBD)
    
    if len(usuarios) > 0:
        cursor = conexionBD.cursor()
        cursor.execute("SELECT COUNT(*) FROM prestamos")
        total_prestamos = cursor.fetchone()[0]

        if total_prestamos > 0:
            input(f"\n{funciones.ROJO}...¡No se puede vaciar! Hay prestamos registrados. Presione Enter...{funciones.RESET}")
        else:
            texto = f"{funciones.ROJO}¿Deseas borrar TODOS los usuarios (Si/No)? {funciones.RESET}\033[s"
            funciones.escritura_lenta_print(texto)
            opc = input("\033[u").lower().strip()

            while opc not in ["si", "no"]:
                print(f"\n{funciones.ROJO}Opcion invalida. Escriba unicamente SI o NO.{funciones.RESET}")
                funciones.espereTecla()
                print("\033[u\033[J", end="", flush=True)
                opc = input().lower().strip()
                    
            if opc == "si":       
                respuesta = Usuarios.crud.vaciar(conexionBD)
                
                if respuesta:
                    funciones.accionExitosa()
                else:
                    funciones.accionNoExitosa()
    else:        
        input(f"\n{funciones.ROJO}...¡No hay usuarios que borrar! Presione Enter...{funciones.RESET}") 

    funciones.menuUsuarios()

def modificarUsuarios(conexionBD):
    print("\033c")
    texto = f"{funciones.AZUL}\n\t\t\t\t...:::: MODIFICAR USUARIOS ::::...\n"
    funciones.escritura_lenta_print(texto)
    
    patron_matricula = r'^\d{1,11}$'

    texto = f"{funciones.AMARILLO}Escribir la matrícula del usuario que quieres actualizar: {funciones.RESET}\033[s"
    funciones.escritura_lenta_print(texto)
    
    matricula = input("\033[u").strip()
    
    while matricula == "" or not re.match(patron_matricula, matricula):
        print(f"\n{funciones.ROJO}Entrada inválida. Ingrese solo números con una longitud máxima de 11 dígitos.{funciones.RESET}")
        funciones.espereTecla()
        print("\033[u\033[J", end="", flush=True)
        matricula = input().strip()
    
    usuarios = Usuarios.crud.buscar(matricula, conexionBD)
    
    if len(usuarios) > 0:
        usuario_actual = usuarios[0] 
        nombre_ant = usuario_actual[1]
        correo_ant = usuario_actual[2]
        carrera_ant = usuario_actual[3]
        cuatri_ant = usuario_actual[4]
        modalidad_ant = usuario_actual[5]

        print(f"\n\t{funciones.AMARILLO}{'Matrícula':<15}{'Nombre':<25}{'Correo':<20}{'Carrera':<25}{'Cuatrimestre':<15}{'Modalidad':<15}{funciones.RESET}\n")
        
        mat = str(usuario_actual[0])[:14]
        nom = str(nombre_ant)[:24]
        cor = str(correo_ant)[:29]
        car = str(carrera_ant)[:29]
        cua = str(cuatri_ant)[:14]
        mod = str(modalidad_ant)[:14]
        
        print(f"\t{mat:<15}{nom:<25}{cor:<20}{car:<25}{cua:<15}{mod:<15}\n")
        
        texto = f"{funciones.ROJO}¿Deseas actualizar a este usuario (Si/No)? {funciones.RESET}\033[s"
        funciones.escritura_lenta_print(texto)
        opc = input("\033[u").lower().strip()
        
        while opc not in ["si", "no"]:
            print(f"\n{funciones.ROJO}Opción inválida. Escriba únicamente SI o NO.{funciones.RESET}")
            funciones.espereTecla()
            print("\033[u\033[J", end="", flush=True)
            opc = input().lower().strip()
            
        if opc == "si":
            print(f"\n{funciones.AZUL}(Presiona ENTER sin escribir nada para conservar el valor actual){funciones.RESET}\n")

            patron_nombre = r'^[A-ZÁÉÍÓÚÑa-záéíóúñ\s]{3,50}$'
            patron_correo = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
            patron_carrera = r'^[A-ZÁÉÍÓÚÑa-záéíóúñ\s]{3,50}$'

            texto = f"{funciones.AMARILLO}Introducir el nuevo nombre: {funciones.RESET}\033[s"
            funciones.escritura_lenta_print(texto)
            entrada = input("\033[u").strip()
            
            while entrada != "" and not re.match(patron_nombre, entrada):
                print(f"\n{funciones.ROJO}Nombre inválido. Solo letras y espacios con una longitud entre 3 y 50 caracteres.{funciones.RESET}")
                funciones.espereTecla()
                print("\033[u\033[J", end="", flush=True)
                entrada = input().strip()
                
            name = nombre_ant if not entrada else entrada.upper()

            texto = f"{funciones.AMARILLO}Introducir el nuevo correo electrónico: {funciones.RESET}\033[s"
            funciones.escritura_lenta_print(texto)
            entrada = input("\033[u").strip()
            
            while entrada != "" and not re.match(patron_correo, entrada):
                print(f"\n{funciones.ROJO}Formato de correo inválido, {funciones.VERDE}Ej: usuario@dominio.com{funciones.RESET}")
                funciones.espereTecla()
                print("\033[u\033[J", end="", flush=True)
                entrada = input().strip()
                
            correo = correo_ant if not entrada else entrada.lower()

            texto = f"{funciones.AMARILLO}Introducir la nueva carrera: {funciones.RESET}\033[s"
            funciones.escritura_lenta_print(texto)
            entrada = input("\033[u").strip()

            while entrada != "" and not re.match(patron_carrera, entrada):
                print(f"\n{funciones.ROJO}Carrera inválida. Solo letras y espacios con una longitud entre 3 y 50 caracteres.{funciones.RESET}")
                funciones.espereTecla()
                print("\033[u\033[J", end="", flush=True)
                entrada = input().strip()

            carrera = carrera_ant if not entrada else entrada.upper()

            texto = f"{funciones.AMARILLO}Seleccionar el nuevo cuatrimestre: {funciones.RESET}\033[s"
            funciones.escritura_lenta_print(texto)
            entrada = input("\033[u").strip()

            while entrada != "" and entrada not in [str(i) for i in range(1, 12)]:
                print(f"\n{funciones.ROJO}Cuatrimestre invalido. Seleccione un numero del 1 al 11.{funciones.RESET}")
                funciones.espereTecla()
                print("\033[u\033[J", end="", flush=True)
                entrada = input().strip()

            cuatrimestre = cuatri_ant if not entrada else entrada

            texto = f"{funciones.AMARILLO}Seleccionar la nueva modalidad (1 = BIS, 2 = CLASICA): {funciones.RESET}\033[s"
            funciones.escritura_lenta_print(texto)
            entrada = input("\033[u").strip()

            while entrada != "" and entrada not in ["1", "2"]:
                print(f"\n{funciones.ROJO}Modalidad inválida. Seleccione 1 para BIS o 2 para CLASICA.{funciones.RESET}")
                funciones.espereTecla()
                print("\033[u\033[J", end="", flush=True)
                entrada = input().strip()

            if not entrada:
                modalidad = modalidad_ant
            else:
                modalidad = "BIS" if entrada == "1" else "CLASICA"

            respuesta = Usuarios.crud.actualizar(matricula, name, correo, carrera, cuatrimestre, modalidad, conexionBD)
        
            if respuesta:
                funciones.accionExitosa()
                funciones.menuUsuarios()
            else:
                funciones.accionNoExitosa()
                funciones.menuUsuarios()    
        else:
            funciones.menuUsuarios()
    else:        
        input(f"\n{funciones.ROJO}...¡No se encontró ningún usuario con la matrícula: {matricula}! Presione Enter...{funciones.RESET}")
        funciones.menuUsuarios()

def generarReporteExcel(conexionBD):
    print("\033c")

    texto = (
        f"{funciones.AZUL}"
        + "....:::: GENERAR REPORTE DE USUARIOS EN EXCEL ::::....".center(80)
        + f"\n{funciones.RESET}"
    )
    funciones.escritura_lenta_print(texto)

    try:
        lista_usuarios = Usuarios.crud.consultar(conexionBD)

        if len(lista_usuarios) == 0:
            print(
                f"\n{funciones.ROJO}"
                "No hay usuarios registrados para generar el reporte."
                f"{funciones.RESET}"
            )
        else:
            archivo_excel = Workbook()
            hoja = archivo_excel.active
            hoja.title = "Reporte de usuarios"

            hoja.merge_cells("A1:F1")
            hoja["A1"] = "SISTEMA DE BIBLIOTECA UTD"
            hoja["A1"].font = Font(bold=True, size=16)
            hoja["A1"].alignment = Alignment(
                horizontal="center",
                vertical="center"
            )

            hoja.merge_cells("A2:F2")
            hoja["A2"] = "REPORTE GENERAL DE USUARIOS"
            hoja["A2"].font = Font(bold=True, size=14)
            hoja["A2"].alignment = Alignment(
                horizontal="center",
                vertical="center"
            )

            fecha_actual = datetime.now().strftime("%d/%m/%Y %H:%M:%S")

            hoja.merge_cells("A3:F3")
            hoja["A3"] = f"Fecha de generacion: {fecha_actual}"
            hoja["A3"].alignment = Alignment(horizontal="center")

            encabezados = [
                "Matricula",
                "Nombre",
                "Correo",
                "Carrera",
                "Cuatrimestre",
                "Modalidad"
            ]

            relleno_encabezado = PatternFill(
                fill_type="solid",
                fgColor="1F4E78"
            )

            borde_delgado = Side(
                style="thin",
                color="000000"
            )

            borde = Border(
                left=borde_delgado,
                right=borde_delgado,
                top=borde_delgado,
                bottom=borde_delgado
            )

            fila_encabezados = 5

            for columna, encabezado in enumerate(encabezados, start=1):
                celda = hoja.cell(
                    row=fila_encabezados,
                    column=columna,
                    value=encabezado
                )

                celda.font = Font(
                    bold=True,
                    color="FFFFFF"
                )

                celda.fill = relleno_encabezado

                celda.alignment = Alignment(
                    horizontal="center",
                    vertical="center"
                )

                celda.border = borde

            fila = 6
            contador_usuarios = 0

            for usuario in lista_usuarios:
                contador_usuarios += 1

                datos_usuario = {
                    "matricula": str(usuario[0]),
                    "nombre": usuario[1],
                    "correo": usuario[2],
                    "carrera": usuario[3],
                    "cuatrimestre": usuario[4],
                    "modalidad": usuario[5]
                }

                for columna, dato in enumerate(datos_usuario.values(), start=1):
                    celda = hoja.cell(
                        row=fila,
                        column=columna,
                        value=dato
                    )

                    celda.alignment = Alignment(
                        horizontal="center",
                        vertical="center",
                        wrap_text=True
                    )

                    celda.border = borde

                fila += 1

            fila_total = fila + 1

            hoja.cell(
                row=fila_total,
                column=1,
                value="Total de usuarios:"
            ).font = Font(bold=True)

            hoja.cell(
                row=fila_total,
                column=2,
                value=contador_usuarios
            ).font = Font(bold=True)

            hoja.column_dimensions["A"].width = 18
            hoja.column_dimensions["B"].width = 30
            hoja.column_dimensions["C"].width = 35
            hoja.column_dimensions["D"].width = 35
            hoja.column_dimensions["E"].width = 15
            hoja.column_dimensions["F"].width = 18

            hoja.row_dimensions[1].height = 25
            hoja.row_dimensions[2].height = 23
            hoja.row_dimensions[5].height = 25

            hoja.freeze_panes = "A6"
            hoja.auto_filter.ref = f"A5:F{fila - 1}"

            descargas = Path.home() / "Downloads"
            carpeta_reportes_user = descargas / "Reportes Usuarios"

            if not os.path.exists(carpeta_reportes_user):
                os.makedirs(carpeta_reportes_user)

            fecha_archivo = datetime.now().strftime("%Y%m%d_%H%M%S")

            nombre_archivo = (
                f"Reporte_Usuarios_{fecha_archivo}.xlsx"
            )

            ruta_archivo = os.path.join(
                carpeta_reportes_user,
                nombre_archivo
            )

            archivo_excel.save(ruta_archivo)

            print(
                f"\n{funciones.VERDE}"
                "Reporte de usuarios generado correctamente."
                f"{funciones.RESET}"
            )

            print(
                f"\n{funciones.AMARILLO}"
                f"Total de usuarios exportados: {contador_usuarios}"
                f"{funciones.RESET}"
            )

            print(
                f"\n{funciones.AMARILLO}"
                f"Archivo guardado en:\n{os.path.abspath(ruta_archivo)}"
                f"{funciones.RESET}"
            )

    except PermissionError:
        print(
            f"\n{funciones.ROJO}"
            "No se pudo guardar el reporte porque el archivo esta abierto."
            "\nCierra el archivo de Excel e intentalo nuevamente."
            f"{funciones.RESET}"
        )

    except Exception as error:
        print(
            f"\n{funciones.ROJO}"
            f"No se pudo generar el reporte: {error}"
            f"{funciones.RESET}"
        )

    funciones.espereTecla()
    funciones.menuUsuarios()