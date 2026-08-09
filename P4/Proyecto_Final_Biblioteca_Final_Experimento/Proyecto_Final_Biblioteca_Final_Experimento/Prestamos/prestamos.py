from datetime import datetime
import re
import funciones
import Usuarios.crud
import Prestamos.crud
import Libros.crud
import os
import funciones
import Usuarios.crud
import Prestamos.crud
import Libros.crud
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from pathlib import Path
from datetime import date

def agregarPrestamos(conexionBD):
    print("\033c")
    texto = f"{funciones.AZUL}\n" + "....:::: REGISTRAR PRÉSTAMO ::::....".center(80) + f"\n{funciones.RESET}"
    funciones.escritura_lenta_print(texto)
    
    patron_matricula = r'^\d{1,11}$'       
    patron_codigo_libro = r'^\d{1,11}$'    
    patron_fecha = r'^\d{4}-\d{2}-\d{2}$'  

    texto = f"{funciones.AMARILLO}Matrícula del Alumno: {funciones.RESET}\033[s"
    funciones.escritura_lenta_print(texto)
    id_usuario = input("\033[u").strip()

    usuario_valido = False
    while not usuario_valido:
        if id_usuario == "":
            print(f"\n{funciones.ROJO}La matrícula no puede estar vacía.{funciones.RESET}")
        elif not re.match(patron_matricula, id_usuario):
            print(f"\n{funciones.ROJO}Matrícula inválida. Ingrese solo números.{funciones.RESET}")
        else:
            usr_encontrado = Usuarios.crud.buscar(id_usuario, conexionBD)
            if len(usr_encontrado) > 0:
                usuario_valido = True
            else:
                print(f"\n{funciones.ROJO}El usuario con matrícula: {id_usuario} No existe en el sistema.{funciones.RESET}")

        if not usuario_valido:
            funciones.espereTecla()
            print("\033[u\033[J", end="", flush=True)
            id_usuario = input().strip()

    texto = f"{funciones.AMARILLO}Código del Libro: {funciones.RESET}\033[s"
    funciones.escritura_lenta_print(texto)
    id_libro = input("\033[u").strip()

    libro_valido = False
    while not libro_valido:
        if id_libro == "":
            print(f"\n{funciones.ROJO}El código del libro no puede estar vacío.{funciones.RESET}")
        elif not re.match(patron_codigo_libro, id_libro):
            print(f"\n{funciones.ROJO}Código inválido. Ingrese solo números.{funciones.RESET}")
        else:
            libro_encontrado = Libros.crud.buscar(id_libro, conexionBD)
            if len(libro_encontrado) > 0:
                libro_valido = True
            else:
                print(f"\n{funciones.ROJO}El libro con el código: {id_libro} No existe en la biblioteca.{funciones.RESET}")

        if not libro_valido:
            funciones.espereTecla()
            print("\033[u\033[J", end="", flush=True)
            id_libro = input().strip()

    texto = f"{funciones.AMARILLO}Fecha de Préstamo (Año (Ej.2026) -Mes (Ej.07) -Dia(Ej.14)): {funciones.RESET}\033[s"
    funciones.escritura_lenta_print(texto)
    fecha_prestamo = input("\033[u").strip()

    obj_fecha_p = None
    fecha_p_valida = False
    fecha_actual = date.today()

    while not fecha_p_valida:
        if fecha_prestamo == "":
            print(f"\n{funciones.ROJO}La fecha de préstamo no puede estar vacía.{funciones.RESET}")
        elif not re.match(patron_fecha, fecha_prestamo):
            print(f"\n{funciones.ROJO}Formato de fecha incorrecto. Debe ser (Año (Ej.2026) -Mes (Ej.07) -Dia(Ej.14)).{funciones.RESET}")
        else:
            try:
                obj_fecha_p = datetime.strptime(fecha_prestamo, "%Y-%m-%d")
                
                if obj_fecha_p.date() < fecha_actual:
                    print(f"\n{funciones.ROJO}La fecha del préstamo no puede ser menor a la actual.{funciones.RESET}")
                else:
                    fecha_p_valida = True
            except ValueError:
                print(f"\n{funciones.ROJO}La fecha ingresada no existe en el calendario.{funciones.RESET}")

        if not fecha_p_valida:
            funciones.espereTecla()
            print("\033[u\033[J", end="", flush=True)
            fecha_prestamo = input().strip()

    texto = f"{funciones.AMARILLO}Fecha de Devolución programada (Año (Ej.2026) -Mes (Ej.07) -Dia(Ej.14)): {funciones.RESET}\033[s"
    funciones.escritura_lenta_print(texto)
    fecha_devolucion = input("\033[u").strip()

    fecha_d_valida = False
    while not fecha_d_valida:
        if fecha_devolucion == "":
            print(f"\n{funciones.ROJO}La fecha de devolución no puede estar vacía.{funciones.RESET}")
        elif not re.match(patron_fecha, fecha_devolucion):
            print(f"\n{funciones.ROJO}Formato de fecha incorrecto. Debe ser (Año (Ej.2026) -Mes (Ej.07) -Dia(Ej.14)).{funciones.RESET}")
        else:
            try:
                obj_fecha_d = datetime.strptime(fecha_devolucion, "%Y-%m-%d")
                if obj_fecha_d < obj_fecha_p:
                    print(f"\n{funciones.ROJO}La fecha de devolución no puede ser anterior a la fecha de préstamo.{funciones.RESET}")
                else:
                    fecha_d_valida = True
            except ValueError:
                print(f"\n{funciones.ROJO}La fecha ingresada no existe.{funciones.RESET}")

        if not fecha_d_valida:
            funciones.espereTecla()
            print("\033[u\033[J", end="", flush=True)
            fecha_devolucion = input().strip()

    estado = "ACTIVO"

    texto = f"{funciones.AMARILLO}Observaciones (Opcional): {funciones.RESET}\033[s"
    funciones.escritura_lenta_print(texto)
    observaciones = input("\033[u").strip()

    while len(observaciones) > 255:
        print(f"\n{funciones.ROJO}Las observaciones superan el límite de 255 caracteres.{funciones.RESET}")
        funciones.espereTecla()
        print("\033[u\033[J", end="", flush=True)
        observaciones = input().strip()

    if observaciones == "":
        observaciones = "Sin observaciones"

    respuesta = Prestamos.crud.insertar(id_usuario, id_libro, fecha_prestamo, fecha_devolucion, estado, observaciones, conexionBD)
    
    if respuesta: 
        funciones.accionExitosa()
    else:
        funciones.accionNoExitosa()
        
    funciones.menuPrestamos()

def mostrarPrestamos(conexionBD): 
    print("\033c")
    texto = f"{funciones.AZUL}" + "....:::: LISTA DE PRÉSTAMOS ::::....".center(80) + f"\n{funciones.RESET}"
    funciones.escritura_lenta_print(texto)
    
    prestamos = Prestamos.crud.consultar(conexionBD)
    
    if len(prestamos) > 0:
        print(f"\t{funciones.AMARILLO}{'ID':<5}\t{'Usuario':<12}\t{'Libro':<10}\t{'F. Préstamo':<12}\t{'F. Devolución':<14}\t{'Estado':<10}\t{'Observaciones':<20}{funciones.RESET}\n")
        
        for p in prestamos:
            usr = str(p[1])[:10]
            libro = str(p[2])[:8]
            estado = str(p[5])[:9]
            obs = str(p[6]) if p[6] else "Sin obs."
            
            print(f"\t{str(p[0]):<5}\t{usr:<12}\t{libro:<10}\t{str(p[3]):<12}\t{str(p[4]):<14}\t{estado:<10}\t{obs:<20}")

            
        print("\n")
        funciones.espereTecla()
        funciones.menuPrestamos()
    else:        
        texto = f"\n" + f"{funciones.ROJO}...¡No hay préstamos que mostrar!...{funciones.RESET}".center(80) + "\n"
        funciones.escritura_lenta_print(texto)
        funciones.espereTecla()
        funciones.menuPrestamos()

def buscarPrestamos(conexionBD):
    print("\033c")
    texto = f"{funciones.AZUL}\n" + "....:::: BUSCAR PRÉSTAMO ::::....".center(80) + f"\n{funciones.RESET}"
    funciones.escritura_lenta_print(texto)
    
    patron_id = r'^\d{1,11}$'
    texto = f"{funciones.AMARILLO}Escribir id del préstamo o matricula del usuario a buscar: {funciones.RESET}\033[s"
    funciones.escritura_lenta_print(texto)
    id_prestamo = input("\033[u").strip()

    while id_prestamo == "" or not re.match(patron_id, id_prestamo):
        if id_prestamo == "":
            print(f"\n{funciones.ROJO}El id o matricula no puede estar vacío.{funciones.RESET}")
        else:
            print(f"\n{funciones.ROJO}Id o matricula inválida. El id o matricula debe ser únicamente números.{funciones.RESET}")
            
        funciones.espereTecla()
        print("\033[u\033[J", end="", flush=True)
        id_prestamo = input().strip()

    prestamos = Prestamos.crud.buscar(id_prestamo, conexionBD)
    
    if len(prestamos) > 0:
        cont_prestamos=0
        for p in prestamos:
            cont_prestamos+=1
        print(f"{funciones.VERDE}Se encontraron {cont_prestamos} registros{funciones.RESET}\n")
        print(f"")
        print(f"\n\t{funciones.AMARILLO}{'ID':<5}\t{'Matrícula':<12}\t{'Libro':<10}\t{'F. Préstamo':<12}\t{'F. Devolución':<14}\t{'Estado':<10}\t{'Observaciones':<20}{funciones.RESET}\n")
        
        for p in prestamos:
            usr = str(p[1])[:10]
            libro = str(p[2])[:8]
            estado = str(p[5])[:9]
            obs = str(p[6]) if p[6] else "Sin obs."
            
            print(f"\t{str(p[0]):<5}\t{usr:<12}\t{libro:<10}\t{str(p[3]):<12}\t{str(p[4]):<14}\t{estado:<10}\t{obs:<20}")
            
        print("\n")
        funciones.espereTecla()
        funciones.menuPrestamos()
    else:        
        texto = f"\n" + f"{funciones.ROJO}...¡No se encontró ningún préstamo con el id/matricula:{id_prestamo}!...{funciones.RESET}".center(80) + "\n"
        funciones.escritura_lenta_print(texto)
        funciones.espereTecla()
        funciones.menuPrestamos()

def borrarPrestamos(conexionBD):
    print("\033c")
    texto = f"{funciones.AZUL}\n" + "....:::: ELIMINAR PRÉSTAMO ::::....".center(80) + f"\n{funciones.RESET}"
    funciones.escritura_lenta_print(texto)
    
    patron_id = r'^\d{1,11}$'

    texto = f"{funciones.AMARILLO}Escribir el id del préstamo a borrar: {funciones.RESET}\033[s"
    funciones.escritura_lenta_print(texto)
    id_prestamo = input("\033[u").strip()

    while id_prestamo == "" or not re.match(patron_id, id_prestamo):
        if id_prestamo == "":
            print(f"\n{funciones.ROJO}El id del préstamo no puede estar vacío.{funciones.RESET}")
        else:
            print(f"\n{funciones.ROJO}Id inválido. Ingrese solo números.{funciones.RESET}")
            
        funciones.espereTecla()
        print("\033[u\033[J", end="", flush=True)
        id_prestamo = input().strip()
        
    prestamos = Prestamos.crud.buscarID(id_prestamo, conexionBD)
    
    if len(prestamos) > 0:
        p = prestamos[0]
        obs = str(p[6]) if p[6] else "Sin obs."
        
        print(f"\n\t{funciones.AMARILLO}{'ID':<5}\t{'Matrícula':<12}\t{'Libro':<10}\t{'F. Préstamo':<12}\t{'F. Devolución':<14}\t{'Estado':<10}\t{'Observaciones':<20}{funciones.RESET}\n")
        print(f"\t{str(p[0]):<5}\t{str(p[1])[:10]:<12}\t{str(p[2])[:8]:<10}\t{str(p[3]):<12}\t{str(p[4]):<14}\t{str(p[5])[:9]:<10}\t{obs:<20}\n")
        
        texto = f"{funciones.ROJO}¿Deseas borrar este préstamo (Si/No)? {funciones.RESET}\033[s"
        funciones.escritura_lenta_print(texto)
        
        opc = input("\033[u").lower().strip()
        
        while opc != "si" and opc != "no":
            print(f"\n{funciones.ROJO}Opción inválida. Escriba unicamente SI o NO.{funciones.RESET}")
            funciones.espereTecla()
            
            print("\033[u\033[J", end="", flush=True)
            opc = input().lower().strip()
                
        if opc == "si":       
            respuesta = Prestamos.crud.borrar(id_prestamo, conexionBD)
            if respuesta:
                funciones.accionExitosa()
                funciones.menuPrestamos()
            else:
                funciones.accionNoExitosa()
                funciones.menuPrestamos() 
        else:
            funciones.menuPrestamos()
    else:        
        input(f"\n{funciones.ROJO}...¡No se encontró ningún préstamo con el id: {id_prestamo}! Presione Enter...{funciones.RESET}")
        funciones.menuPrestamos()

def limpiarPrestamos(conexionBD):
    print("\033c")
    texto = f"{funciones.AZUL}\n" + "....:::: VACIAR HISTORIAL DE PRÉSTAMOS ::::....".center(80) + f"\n{funciones.RESET}"
    funciones.escritura_lenta_print(texto)
    
    prestamos = Prestamos.crud.consultar(conexionBD)
    
    if len(prestamos) > 0:
        texto = f"\n{funciones.ROJO}¿Deseas vaciar TODOS los préstamos del sistema (Si/No)? {funciones.RESET}\033[s"
        funciones.escritura_lenta_print(texto)
        
        opc = input("\033[u").lower().strip()
        
        while opc != "si" and opc != "no":
            print(f"\n{funciones.ROJO}Opción inválida. Escriba unicamente SI o NO.{funciones.RESET}")
            funciones.espereTecla()
            
            print("\033[u\033[J", end="", flush=True)
            opc = input().lower().strip()
                
        if opc == "si":       
            respuesta = Prestamos.crud.vaciar(conexionBD)
            if respuesta:
                funciones.accionExitosa()
                funciones.menuPrestamos()
            else:
                funciones.accionNoExitosa()
                funciones.menuPrestamos()      
        else:
            funciones.menuPrestamos()
    else:        
        input(f"\n{funciones.ROJO}...¡No hay préstamos que borrar! Presione Enter...{funciones.RESET}") 
        funciones.menuPrestamos()

def modificarPrestamos(conexionBD):
    print("\033c")
    texto = f"{funciones.AZUL}\n" + "....:::: MODIFICAR UN PRÉSTAMO ::::....".center(80) + f"\n{funciones.RESET}"
    funciones.escritura_lenta_print(texto)
    
    patron_id = r'^\d{1,11}$'
    patron_matricula = r'^\d{1,11}$' 
    patron_fecha = r'^\d{4}-\d{2}-\d{2}$'

    texto = f"{funciones.AMARILLO}Escribir el id del préstamo que quieres actualizar: {funciones.RESET}\033[s"
    funciones.escritura_lenta_print(texto)
    id_prestamo = input("\033[u").strip()

    while id_prestamo == "" or not re.match(patron_id, id_prestamo):
        if id_prestamo == "":
            print(f"\n{funciones.ROJO}El ID no puede estar vacío.{funciones.RESET}")
        else:
            print(f"\n{funciones.ROJO}Entrada inválida. Ingrese solo los numeros del id.{funciones.RESET}")
            
        funciones.espereTecla()
        print("\033[u\033[J", end="", flush=True)
        id_prestamo = input().strip()

    prestamos = Prestamos.crud.buscarID(id_prestamo, conexionBD)

    if len(prestamos) > 0:
        p_act = prestamos[0]
        id_usuario_ant = p_act[1]
        id_libro_ant = p_act[2]
        fecha_p_ant = str(p_act[3])
        fecha_d_ant = str(p_act[4])
        estado_ant = p_act[5]
        obs_ant = p_act[6] if p_act[6] else "Sin observaciones"

        print(f"\n\t{funciones.AMARILLO}--- DATOS ACTUALES DEL PRÉSTAMO #{id_prestamo} ---{funciones.RESET}")
        print(f"\tUsuario (Matrícula) : {id_usuario_ant}")
        print(f"\tLibro (Código)      : {id_libro_ant}")
        print(f"\tF. Préstamo / Dev.  : {fecha_p_ant}  |  {fecha_d_ant}")
        print(f"\tEstado / Obs.       : {estado_ant}  |  {obs_ant}\n")
        
        texto = f"{funciones.ROJO}¿Deseas actualizar este préstamo (Si/No)? {funciones.RESET}\033[s"
        funciones.escritura_lenta_print(texto)
        opc = input("\033[u").lower().strip()
        
        while opc not in ["si", "no"]:
            print(f"\n{funciones.ROJO}Opción inválida. Escriba unicamente SI o NO.{funciones.RESET}")
            funciones.espereTecla()
            print("\033[u\033[J", end="", flush=True)
            opc = input().lower().strip()

        if opc == "si":
            print(f"\n{funciones.AZUL}(Presiona ENTER sin escribir nada para conservar el valor actual){funciones.RESET}\n")

            texto = f"{funciones.AMARILLO}Nuevo ID Usuario [{id_usuario_ant}]: {funciones.RESET}\033[s"
            funciones.escritura_lenta_print(texto)
            entrada = input("\033[u").strip()
            
            usuario_valido = False
            while not usuario_valido:
                if entrada == "":
                    id_usuario = id_usuario_ant
                    usuario_valido = True
                elif not re.match(patron_matricula, entrada):
                    print(f"\n{funciones.ROJO}Matrícula inválida. Ingrese solo numeros.{funciones.RESET}")
                else:
                    usr_encontrado = Usuarios.crud.buscarAlumno(entrada, conexionBD)
                    if len(usr_encontrado) > 0:
                        id_usuario = entrada
                        usuario_valido = True
                    else:
                        print(f"\n{funciones.ROJO}El usuario con matrícula: {entrada} No existe en la base de datos.{funciones.RESET}")

                if not usuario_valido:
                    funciones.espereTecla()
                    print("\033[u\033[J", end="", flush=True)
                    entrada = input().strip()

            texto = f"{funciones.AMARILLO}Nuevo Código Libro [{id_libro_ant}]: {funciones.RESET}\033[s"
            funciones.escritura_lenta_print(texto)
            entrada = input("\033[u").strip()

            libro_valido = False
            while not libro_valido:
                if entrada == "":
                    id_libro = id_libro_ant
                    libro_valido = True
                elif not re.match(patron_id, entrada):
                    print(f"\n{funciones.ROJO}Código inválido. Ingrese solo números.{funciones.RESET}")
                else:
                    libro_encontrado = Libros.crud.buscar(entrada, conexionBD)
                    if len(libro_encontrado) > 0:
                        id_libro = entrada
                        libro_valido = True
                    else:
                        print(f"\n{funciones.ROJO}El libro con código: {entrada} No existe en el catálogo.{funciones.RESET}")

                if not libro_valido:
                    funciones.espereTecla()
                    print("\033[u\033[J", end="", flush=True)
                    entrada = input().strip()

            texto = f"{funciones.AMARILLO}Nueva Fecha Préstamo [{fecha_p_ant}]: {funciones.RESET}\033[s"
            funciones.escritura_lenta_print(texto)
            entrada = input("\033[u").strip()
            
            fecha_p_valida = False
            while not fecha_p_valida:
                if entrada == "":
                    fecha_prestamo = fecha_p_ant
                    fecha_p_valida = True
                elif not re.match(patron_fecha, entrada):
                    print(f"\n{funciones.ROJO}Formato de fecha incorrecto (Año (Ej.2026) -Mes (Ej.07) -Dia(Ej.14)).{funciones.RESET}")
                else:
                    try:
                        datetime.strptime(entrada, "%Y-%m-%d")
                        fecha_prestamo = entrada
                        fecha_p_valida = True
                    except ValueError:
                        print(f"\n{funciones.ROJO}La fecha no existe.{funciones.RESET}")

                if not fecha_p_valida:
                    funciones.espereTecla()
                    print("\033[u\033[J", end="", flush=True)
                    entrada = input().strip()

            texto = f"{funciones.AMARILLO}Nueva Fecha Devolución [{fecha_d_ant}]: {funciones.RESET}\033[s"
            funciones.escritura_lenta_print(texto)
            entrada = input("\033[u").strip()

            fecha_d_valida = False
            while not fecha_d_valida:
                if entrada == "":
                    fecha_devolucion = fecha_d_ant
                    fecha_d_valida = True
                elif not re.match(patron_fecha, entrada):
                    print(f"\n{funciones.ROJO}Formato de fecha incorrecto (Año (Ej.2026) -Mes (Ej.07) -Dia(Ej.14)).{funciones.RESET}")
                else:
                    try:
                        obj_p = datetime.strptime(fecha_prestamo, "%Y-%m-%d")
                        obj_d = datetime.strptime(entrada, "%Y-%m-%d")
                        if obj_d < obj_p:
                            print(f"\n{funciones.ROJO}La devolución no puede ser anterior al préstamo.{funciones.RESET}")
                        else:
                            fecha_devolucion = entrada
                            fecha_d_valida = True
                    except ValueError:
                        print(f"\n{funciones.ROJO}La fecha no existe.{funciones.RESET}")

                if not fecha_d_valida:
                    funciones.espereTecla()
                    print("\033[u\033[J", end="", flush=True)
                    entrada = input().strip()

            texto = f"{funciones.AMARILLO}Nuevo Estado [{estado_ant}] (ACTIVO/ENTREGADO): {funciones.RESET}\033[s"
            funciones.escritura_lenta_print(texto)
            entrada = input("\033[u").upper().strip()
            
            estados_validos = ["ACTIVO", "ENTREGADO"]
            while entrada != "" and entrada not in estados_validos:
                print(f"\n{funciones.ROJO}Estado inválido. Opciones válidas: {', '.join(estados_validos)}.{funciones.RESET}")
                funciones.espereTecla()
                print("\033[u\033[J", end="", flush=True)
                entrada = input().upper().strip()
                
            estado = estado_ant if entrada == "" else entrada

            texto = f"{funciones.AMARILLO}Nueva Observación [{obs_ant}]: {funciones.RESET}\033[s"
            funciones.escritura_lenta_print(texto)
            entrada = input("\033[u").strip()

            while len(entrada) > 255:
                print(f"\n{funciones.ROJO}La observación no debe exceder 255 caracteres.{funciones.RESET}")
                funciones.espereTecla()
                print("\033[u\033[J", end="", flush=True)
                entrada = input().strip()
                
            observaciones = obs_ant if entrada == "" else entrada

            respuesta = Prestamos.crud.actualizar(
                id_prestamo, id_usuario, id_libro, fecha_prestamo, 
                fecha_devolucion, estado, observaciones, conexionBD
            )
        
            if respuesta:
                funciones.accionExitosa()
                funciones.menuPrestamos()
            else:
                funciones.accionNoExitosa()
                funciones.menuPrestamos()    
        else:
            funciones.menuPrestamos()
    else:        
        input(f"\n{funciones.ROJO}...¡No se encontró ningún préstamo con el id: {id_prestamo} presione ENTER{funciones.RESET}")
        funciones.menuPrestamos()


def agregarPrestamosAlumno(matricula,conexionBD):
    print("\033c")
    texto = f"{funciones.AZUL}\n" + "....:::: REGISTRAR PRÉSTAMO ::::....".center(80) + f"\n{funciones.RESET}"
    funciones.escritura_lenta_print(texto)
        
    patron_codigo_libro = r'^\d{1,11}$'    
    patron_fecha = r'^\d{4}-\d{2}-\d{2}$'  

    texto = f"{funciones.AMARILLO}Código del Libro: {funciones.RESET}\033[s"
    funciones.escritura_lenta_print(texto)
    id_libro = input("\033[u").strip()

    libro_valido = False
    while not libro_valido:
        if id_libro == "":
            print(f"\n{funciones.ROJO}El código del libro no puede estar vacío.{funciones.RESET}")
        elif not re.match(patron_codigo_libro, id_libro):
            print(f"\n{funciones.ROJO}Código inválido. Ingrese solo números.{funciones.RESET}")
        else:
            libro_encontrado = Libros.crud.buscar(id_libro, conexionBD)
            if len(libro_encontrado) > 0:
                libro_valido = True
            else:
                print(f"\n{funciones.ROJO}El libro con el código: {id_libro} No existe en la biblioteca.{funciones.RESET}")

        if not libro_valido:
            funciones.espereTecla()
            print("\033[u\033[J", end="", flush=True)
            id_libro = input().strip()

    texto = f"{funciones.AMARILLO}Fecha de Préstamo (Año (Ej.2026) -Mes (Ej.07) -Dia(Ej.14)): {funciones.RESET}\033[s"
    funciones.escritura_lenta_print(texto)
    fecha_prestamo = input("\033[u").strip()

    obj_fecha_p = None
    fecha_p_valida = False

    while not fecha_p_valida:
        if fecha_prestamo == "":
            print(f"\n{funciones.ROJO}La fecha de préstamo no puede estar vacía.{funciones.RESET}")
        elif not re.match(patron_fecha, fecha_prestamo):
            print(f"\n{funciones.ROJO}Formato de fecha incorrecto. Debe ser (Año (Ej.2026) -Mes (Ej.07) -Dia(Ej.14)).{funciones.RESET}")
        else:
            try:
                obj_fecha_p = datetime.strptime(fecha_prestamo, "%Y-%m-%d")
                fecha_p_valida = True
            except ValueError:
                print(f"\n{funciones.ROJO}La fecha ingresada no existe.{funciones.RESET}")

        if not fecha_p_valida:
            funciones.espereTecla()
            print("\033[u\033[J", end="", flush=True)
            fecha_prestamo = input().strip()

    texto = f"{funciones.AMARILLO}Fecha de Devolución programada (Año (Ej.2026) -Mes (Ej.07) -Dia(Ej.14)): {funciones.RESET}\033[s"
    funciones.escritura_lenta_print(texto)
    fecha_devolucion = input("\033[u").strip()

    fecha_d_valida = False

    while not fecha_d_valida:
        if fecha_devolucion == "":
            print(f"\n{funciones.ROJO}La fecha de devolución no puede estar vacía.{funciones.RESET}")
        elif not re.match(patron_fecha, fecha_devolucion):
            print(f"\n{funciones.ROJO}Formato de fecha incorrecto. Debe ser (Año (Ej.2026) -Mes (Ej.07) -Dia(Ej.14)).{funciones.RESET}")
        else:
            try:
                obj_fecha_d = datetime.strptime(fecha_devolucion, "%Y-%m-%d")
                if obj_fecha_d < obj_fecha_p:
                    print(f"\n{funciones.ROJO}La fecha de devolución no puede ser anterior a la fecha de préstamo.{funciones.RESET}")
                else:
                    fecha_d_valida = True
            except ValueError:
                print(f"\n{funciones.ROJO}La fecha ingresada no existe.{funciones.RESET}")

        if not fecha_d_valida:
            funciones.espereTecla()
            print("\033[u\033[J", end="", flush=True)
            fecha_devolucion = input().strip()

    estado ="ACTIVO"   

    texto = f"{funciones.AMARILLO}Observaciones (Opcional): {funciones.RESET}\033[s"
    funciones.escritura_lenta_print(texto)
    observaciones = input("\033[u").strip()
    while len(observaciones) > 255:
        print(f"\n{funciones.ROJO}Las notas superan el límite de 255 caracteres.{funciones.RESET}")
        funciones.espereTecla()
        print("\033[u\033[J", end="", flush=True)
        observaciones = input().strip()

    if observaciones == "":
        observaciones = "Sin notas"

    respuesta = Prestamos.crud.insertar(matricula, id_libro, fecha_prestamo, fecha_devolucion, estado, observaciones, conexionBD)
    
    if respuesta: 
        funciones.accionExitosa()
        funciones.menuPrincipalAlumno()
    else:
        funciones.accionNoExitosa()
        funciones.menuPrincipalAlumno()

def generarReporteExcel(conexionBD):
    print("\033c")

    texto = (
        f"{funciones.AZUL}"
        + "....:::: GENERAR REPORTE DE PRÉSTAMOS EN EXCEL ::::....".center(80)
        + f"\n{funciones.RESET}"
    )
    funciones.escritura_lenta_print(texto)

    try:
        lista_prestamos = Prestamos.crud.consultar(conexionBD)

        if len(lista_prestamos) == 0:
            print(
                f"\n{funciones.ROJO}"
                "No hay préstamos registrados para generar el reporte."
                f"{funciones.RESET}"
            )
            funciones.espereTecla()
            funciones.menuPrestamos()
            return

        archivo_excel = Workbook()
        hoja = archivo_excel.active
        hoja.title = "Reporte de préstamos"

        hoja.merge_cells("A1:G1")
        hoja["A1"] = "SISTEMA DE BIBLIOTECA UTD"
        hoja["A1"].font = Font(bold=True, size=16)
        hoja["A1"].alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

        hoja.merge_cells("A2:G2")
        hoja["A2"] = "REPORTE GENERAL DE PRÉSTAMOS"
        hoja["A2"].font = Font(bold=True, size=14)
        hoja["A2"].alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

        fecha_actual = datetime.now().strftime("%d/%m/%Y %H:%M:%S")

        hoja.merge_cells("A3:G3")
        hoja["A3"] = f"Fecha de generación: {fecha_actual}"
        hoja["A3"].alignment = Alignment(horizontal="center")

        encabezados = [
            "ID préstamo",
            "Usuario",
            "Libro",
            "Fecha de préstamo",
            "Fecha de devolución",
            "Estado",
            "Observaciones"
        ]

        relleno_encabezado = PatternFill(
            fill_type="solid",
            fgColor="1F4E78"
        )

        borde_delgado = Side(
            style="thin",
            color="000000"
        )

        borde = Border(
            left=borde_delgado,
            right=borde_delgado,
            top=borde_delgado,
            bottom=borde_delgado
        )

        fila_encabezados = 5

        for columna, encabezado in enumerate(encabezados, start=1):
            celda = hoja.cell(
                row=fila_encabezados,
                column=columna,
                value=encabezado
            )

            celda.font = Font(
                bold=True,
                color="FFFFFF"
            )

            celda.fill = relleno_encabezado

            celda.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True
            )

            celda.border = borde

        fila = 6
        contador_prestamos = 0
        
        conteo_estados = {
            "ACTIVO": 0,
            "ENTREGADO": 0
        }

        for prestamo in lista_prestamos:
            contador_prestamos += 1
            estado = str(prestamo[5]).upper().strip()
            
            if estado in conteo_estados:
                conteo_estados[estado] += 1

            fecha_prestamo = prestamo[3]
            fecha_devolucion = prestamo[4]

            if hasattr(fecha_prestamo, "strftime"):
                fecha_prestamo = fecha_prestamo.strftime("%d/%m/%Y")

            if hasattr(fecha_devolucion, "strftime"):
                fecha_devolucion = fecha_devolucion.strftime("%d/%m/%Y")

            datos = [
                prestamo[0],
                prestamo[1],
                prestamo[2],
                fecha_prestamo,
                fecha_devolucion,
                prestamo[5],
                prestamo[6]
            ]

            for columna, dato in enumerate(datos, start=1):
                celda = hoja.cell(
                    row=fila,
                    column=columna,
                    value=dato
                )

                celda.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True
                )

                celda.border = borde

            fila += 1

        fila_resumen = fila + 1

        hoja.cell(
            row=fila_resumen,
            column=1,
            value="Total de préstamos:"
        ).font = Font(bold=True)

        hoja.cell(
            row=fila_resumen,
            column=2,
            value=contador_prestamos
        ).font = Font(bold=True)

        hoja.cell(
            row=fila_resumen + 1,
            column=1,
            value="Préstamos activos:"
        ).font = Font(bold=True)

        hoja.cell(
            row=fila_resumen + 1,
            column=2,
            value=conteo_estados["ACTIVO"]
        ).font = Font(bold=True)

        hoja.cell(
            row=fila_resumen + 2,
            column=1,
            value="Préstamos entregados:"
        ).font = Font(bold=True)

        hoja.cell(
            row=fila_resumen + 2,
            column=2,
            value=conteo_estados["ENTREGADO"]
        ).font = Font(bold=True)

        anchos_columnas = {
            "A": 15,
            "B": 30,
            "C": 35,
            "D": 20,
            "E": 20,
            "F": 18,
            "G": 45
        }

        for columna, ancho in anchos_columnas.items():
            hoja.column_dimensions[columna].width = ancho

        hoja.row_dimensions[1].height = 25
        hoja.row_dimensions[2].height = 23
        hoja.row_dimensions[5].height = 35

        hoja.freeze_panes = "A6"
        hoja.auto_filter.ref = f"A5:G{fila - 1}"

        descargas = Path.home()/"Downloads"
        carpeta_reportes_pres = descargas / "Reportes Prestamos"
      
        if not os.path.exists(carpeta_reportes_pres):
            os.makedirs(carpeta_reportes_pres)

        fecha_archivo = datetime.now().strftime("%Y%m%d_%H%M%S")

        nombre_archivo = (
            f"Reporte_Prestamos_{fecha_archivo}.xlsx"
        )

        ruta_archivo = os.path.join(
            carpeta_reportes_pres,
            nombre_archivo
        )

        archivo_excel.save(ruta_archivo)

        print(
            f"\n{funciones.VERDE}"
            "Reporte de préstamos generado correctamente."
            f"{funciones.RESET}"
        )

        print(
            f"\n{funciones.AMARILLO}"
            f"Total de préstamos exportados: {contador_prestamos}"
            f"{funciones.RESET}"
        )

        print(
            f"\n{funciones.AMARILLO}"
            f"Archivo guardado en:\n{os.path.abspath(ruta_archivo)}"
            f"{funciones.RESET}"
        )

    except PermissionError:
        print(
            f"\n{funciones.ROJO}"
            "No se pudo guardar el reporte porque el archivo está abierto."
            "\nCierra el archivo de Excel e inténtalo nuevamente."
            f"{funciones.RESET}"
        )

    except Exception as error:
        print(
            f"\n{funciones.ROJO}"
            f"No se pudo generar el reporte: {error}"
            f"{funciones.RESET}"
        )

    funciones.espereTecla()
    funciones.menuPrestamos()