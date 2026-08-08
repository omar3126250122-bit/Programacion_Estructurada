from Libros import crud
import funciones
import re
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path

def agregarLibros(conexionBD):
    print("\033c")
    texto = f"\n\t\t{funciones.AZUL}...:::: AGREGAR LIBROS ::::...\n"
    funciones.escritura_lenta_print(texto)

    patron_titulo = r'^[A-ZÁÉÍÓÚÑa-záéíóúñ0-9\s:.,\-#¿?¡!&"\'()$%+]{2,100}$'
    patron_texto = r'^[A-ZÁÉÍÓÚÑa-záéíóúñ\s.]{2,50}$'
    patron_editorial = r'^[A-ZÁÉÍÓÚÑa-záéíóúñ0-9\s.]{2,50}$'
    texto = f"{funciones.AMARILLO}Introducir el nombre del libro: \033[s"
    funciones.escritura_lenta_print(texto)

    
    libro = input("\033[u").strip()
    
    while libro == "" or not re.match(patron_titulo, libro):
        if libro == "":
            print(f"\n{funciones.ROJO}El nombre del libro no puede estar vacío.{funciones.RESET}")
        else:
            print(f"\n{funciones.ROJO}Título inválido.{funciones.RESET}")
            
        funciones.espereTecla()
        print("\033[u\033[J", end="", flush=True)
        libro = input("\033[u").strip()

    libro = libro.upper()

    texto = f"{funciones.AMARILLO}Introducir el nombre del autor: \033[s"
    funciones.escritura_lenta_print(texto)

    
    autor = input("\033[u").strip()
    
    while autor == "" or not re.match(patron_texto, autor):
        if autor == "":
            print(f"\n{funciones.ROJO}El nombre del autor no puede estar vacío.{funciones.RESET}")
        else:
            print(f"\n{funciones.ROJO}Autor inválido, solo debe contener letras.{funciones.RESET}")
            
        funciones.espereTecla()
        print("\033[u\033[J", end="", flush=True)
        autor = input("\033[u").strip()

    autor = autor.upper()

    texto = f"{funciones.AMARILLO}Introducir el nombre de la editorial: \033[s"
    funciones.escritura_lenta_print(texto)

    editorial = input("\033[u").strip()
    
    while editorial == "" or not re.match(patron_editorial, editorial):
        if editorial == "":
            print(f"\n{funciones.ROJO}El nombre de la editorial no puede estar vacío.{funciones.RESET}")
        else:
            print(f"\n{funciones.ROJO}Editorial inválida, Solo debe contener letras, espacios y numeros{funciones.RESET}")
            
        funciones.espereTecla()
        print("\033[u\033[J", end="", flush=True)
        editorial = input("\033[u").strip()

    editorial = editorial.upper()

    texto = f"{funciones.AMARILLO}Introducir el idioma del libro: \033[s"
    funciones.escritura_lenta_print(texto)
    
    idioma = input("\033[u").strip()
    
    while idioma == "" or not re.match(patron_texto, idioma):
        if idioma == "":
            print(f"\n{funciones.ROJO}El idioma no puede estar vacío.{funciones.RESET}")
        else:
            print(f"\n{funciones.ROJO}Idioma inválido. Solo debe contener letras{funciones.RESET}")
            
        funciones.espereTecla()
        print("\033[u\033[J", end="", flush=True)
        idioma = input("\033[u").strip()

    idioma = idioma.upper()
    respuesta = crud.insertar(libro, autor, editorial, idioma, conexionBD)
    if respuesta: 
        funciones.accionExitosa()
        funciones.menuLibros()
    else:
        funciones.accionNoExitosa()
        funciones.menuLibros()

def mostrarLibros(conexionBD): 
    print("\033c")
    texto = f"{funciones.AZUL}" + "....:::: MOSTRAR LIBROS ::::....".center(80) + f"\n{funciones.RESET}"
    funciones.escritura_lenta_print(texto)
    
    libros = crud.consultar(conexionBD)
    
    if len(libros) > 0:
        print(f"\t{funciones.AMARILLO}{'Código':<8}\t{'Título':<30}\t{'Autor':<25}\t{'Editorial':<20}\t{'Idioma':<12}{funciones.RESET}\n")
        
        for i in libros:
            print(f"\t{str(i[0]):<8}\t{str(i[1]):<30}\t{str(i[2]):<25}\t{str(i[3]):<20}\t{str(i[4]):<12}")
            
        print("\n")
        funciones.espereTecla()
        funciones.menuLibros()
    else:        
        texto = f"\n" + f"{funciones.ROJO}...¡No hay libros que mostrar!...{funciones.RESET}".center(80) + "\n"
        funciones.escritura_lenta_print(texto)
        funciones.espereTecla()
        funciones.menuLibros()

def buscarLibros(conexionBD):
    print("\033c")
    texto = f"{funciones.AZUL}\n\t\t\t\t...:::: BUSCAR LIBROS ::::...\n"
    funciones.escritura_lenta_print(texto)
    
    patron_criterio = r'^[A-ZÁÉÍÓÚÑa-záéíóúñ0-9\s:.,\-#¿?¡!&"\'()$%+]{1,100}$'

    texto = f"{funciones.AMARILLO}Escribir el código o el nombre del libro: \033[s"
    funciones.escritura_lenta_print(texto)
    
    criterio = input("\033[u").strip()

    while criterio == "" or not re.match(patron_criterio, criterio):
        if criterio == "":
            print(f"\n{funciones.ROJO}El valor de búsqueda no puede estar vacío.{funciones.RESET}")
        else:
            print(f"\n{funciones.ROJO}Valor inválido. Ingrese un código o nombre válido.{funciones.RESET}")
            
        funciones.espereTecla()
        print("\033[u\033[J", end="", flush=True)
        criterio = input("\033[u").strip()

    criterio = criterio.upper()

    libros = crud.buscar(criterio, conexionBD)
    
    if len(libros) > 0:
        cont_libros=0
        for l in libros:
            cont_libros+=1
        print(f"{funciones.VERDE}Se encontraron {cont_libros} registros{funciones.RESET}\n")
        print(f"\t{funciones.AMARILLO}{'Código':<8}\t{'Título':<30}\t{'Autor':<25}\t{'Editorial':<20}\t{'Idioma':<12}{funciones.RESET}\n")
        for i in libros:
            print(f"\t{str(i[0]):<8}\t{str(i[1]):<30}\t{str(i[2]):<25}\t{str(i[3]):<20}\t{str(i[4]):<12}")
        funciones.espereTecla()
        funciones.menuLibros()
    else:        
        texto = f"\n{funciones.ROJO}...¡No se encontró ningún libro con ese valor!..."
        funciones.escritura_lenta_print(texto)
        funciones.espereTecla()
        funciones.menuLibros()

def borrarLibros(conexionBD):
    print("\033c")
    texto = f"{funciones.AZUL}\n" + "....:::: BORRAR LIBROS ::::....".center(80) + f"\n{funciones.RESET}"
    funciones.escritura_lenta_print(texto)
    
    patron_codigo = r'^\d{1,11}$'

    texto = f"{funciones.AMARILLO}Escribir el código del libro a borrar: \033[s"
    funciones.escritura_lenta_print(texto)
    
    codigo = input("\033[u").strip()

    while codigo == "" or not re.match(patron_codigo, codigo):
        if codigo == "":
            print(f"\n{funciones.ROJO}El código no puede estar vacío.{funciones.RESET}")
        else:
            print(f"\n{funciones.ROJO}Código inválido. Ingrese solo los dígitos del código.{funciones.RESET}")
            
        funciones.espereTecla()
        print("\033[u\033[J", end="", flush=True)
        codigo = input().strip()

    codigo = codigo.upper()
    libros = crud.buscar(codigo, conexionBD)
    
    if len(libros) > 0:
        cursor = conexionBD.cursor()
        cursor.execute("SELECT COUNT(*) FROM prestamos WHERE id_libro = %s", (codigo,))
        prestamos_asociados = cursor.fetchone()[0]

        if prestamos_asociados > 0:
            input(f"\n{funciones.ROJO}...¡No se puede borrar! El libro tiene un préstamo activo en el sistema. Presione Enter!..{funciones.RESET}")
            funciones.menuLibros()
            return

        print(f"\n\t{funciones.AMARILLO}{'Código':<8}\t{'Título':<30}\t{'Autor':<25}\t{'Editorial':<20}\t{'Idioma':<12}{funciones.RESET}\n")
        for i in libros:
            print(f"\t{str(i[0]):<8}\t{str(i[1]):<30}\t{str(i[2]):<25}\t{str(i[3]):<20}\t{str(i[4]):<12}")
        
        texto = f"\n{funciones.ROJO}¿Deseas borrar el libro mostrado (Si/No)? {funciones.RESET}\033[s"
        funciones.escritura_lenta_print(texto)
        
        opc = input("\033[u").lower().strip()
        
        while opc != "si" and opc != "no":
            print(f"\n{funciones.ROJO}Opción inválida. Escriba unicamente SI o NO.{funciones.RESET}")
            funciones.espereTecla()
            print("\033[u\033[J", end="", flush=True)
            opc = input().lower().strip()
                
        if opc == "si":       
            respuesta = crud.borrar(codigo, conexionBD)
            
            if respuesta:
                funciones.accionExitosa()
            else:
                funciones.accionNoExitosa()
                
            funciones.menuLibros() 
        else:
            funciones.menuLibros()
    else:        
        input(f"\n{funciones.ROJO}...¡No se encontró ningún libro con el código: {codigo}! Presione Enter...{funciones.RESET}")
        funciones.menuLibros()

def limpiarLibros(conexionBD):
    print("\033c")
    texto = f"{funciones.AZUL}\n\t\t\t\t...:::: LIMPIAR LIBROS ::::...\n{funciones.RESET}"
    funciones.escritura_lenta_print(texto)
    
    libros = crud.consultar(conexionBD)
    
    if len(libros) > 0:
        cursor = conexionBD.cursor()
        cursor.execute("SELECT COUNT(*) FROM prestamos")
        total_prestamos = cursor.fetchone()[0]

        if total_prestamos > 0:
            input(f"\n{funciones.ROJO}...¡No se puede vaciar! Hay préstamos activos registrados en el sistema. Presione Enter!..{funciones.RESET}")
            funciones.menuLibros()
            
        texto = f"\n{funciones.ROJO}¿Deseas BORRAR TODOS los libros (Si/No)? {funciones.RESET}\033[s"
        funciones.escritura_lenta_print(texto)
                
        opc = input("\033[u").lower().strip() 
        
        while opc != "si" and opc != "no":
            print(f"\n{funciones.ROJO}Opción inválida. Escriba unicamente SI o NO.{funciones.RESET}")
            funciones.espereTecla()
            print("\033[u\033[J", end="", flush=True)
            opc = input().lower().strip()
                        
        if opc == "si":       
            respuesta = crud.vaciar(conexionBD)
            
            if respuesta:
                funciones.accionExitosa()
            else:
                funciones.accionNoExitosa()
                
            funciones.menuLibros() 
        else:
            funciones.menuLibros()
    else:        
        input(f"\n{funciones.ROJO}...¡Por el momento no hay libros que borrar! Presione Enter...{funciones.RESET}")
        funciones.menuLibros()

def modificarLibros(conexionBD):
    print("\033c")
    texto = f"{funciones.AZUL}\n\t\t\t\t...:::: MODIFICAR LIBROS ::::...\n"
    funciones.escritura_lenta_print(texto)
    
    patron_codigo = r'^\d{1,11}$'
    patron_titulo = r'^[A-ZÁÉÍÓÚÑa-záéíóúñ0-9\s:.,\-#¿?¡!&"\'()$%+]{2,100}$'
    patron_texto = r'^[A-ZÁÉÍÓÚÑa-záéíóúñ\s.]{2,50}$'

    texto = f"{funciones.AMARILLO}Escribir el código del libro que quieres actualizar: \033[s"
    funciones.escritura_lenta_print(texto)
    
    codigo = input("\033[u").strip()
    
    while codigo == "" or not re.match(patron_codigo, codigo):
        if codigo == "":
            print(f"\n{funciones.ROJO}El código no puede estar vacío.{funciones.RESET}")
        else:
            print(f"\n{funciones.ROJO}Entrada inválida. Ingrese solo los dígitos del código.{funciones.RESET}")
            
        funciones.espereTecla()
        print("\033[u\033[J", end="", flush=True)
        codigo = input("\033[u").strip()
    
    codigo = codigo.upper()
    libros = crud.buscar(codigo, conexionBD)
    
    if len(libros) > 0:
        libro_actual = libros[0]
        nombre_ant = libro_actual[1]
        autor_ant = libro_actual[2]
        editorial_ant = libro_actual[3]
        idioma_ant = libro_actual[4]

        print(f"\t{funciones.AMARILLO}{'Código':<8}\t{'Título':<30}\t{'Autor':<25}\t{'Editorial':<20}\t{'Idioma':<12}{funciones.RESET}\n")
        for i in libros:
            print(f"\t{str(i[0]):<8}\t{str(i[1]):<30}\t{str(i[2]):<25}\t{str(i[3]):<20}\t{str(i[4]):<12}")

        texto = f"\n{funciones.ROJO}¿Deseas actualizar el libro mostrado (Si/No)? {funciones.RESET}\033[s"
        funciones.escritura_lenta_print(texto)
        opc = input("\033[u").lower().strip()
        while opc != "si" and opc != "no":
            print(f"\n{funciones.ROJO}Opción inválida. Escriba unicamente SI o NO.{funciones.RESET}")
            funciones.espereTecla()
            print("\033[u\033[J", end="", flush=True)
            opc = input().lower().strip()            
        if opc == "si":
                print(f"\n{funciones.AZUL}(Presiona ENTER sin escribir nada para conservar el valor actual){funciones.RESET}\n")

                texto = f"{funciones.AMARILLO}Introducir el nuevo nombre del libro: \033[s"
                funciones.escritura_lenta_print(texto)
                
                entrada = input("\033[u").strip()
                while entrada != "" and not re.match(patron_titulo, entrada):
                    print(f"\n{funciones.ROJO}Título inválido{funciones.RESET}")
                    funciones.espereTecla()
                    print("\033[u\033[J", end="", flush=True)
                    entrada = input("\033[u").strip()
                    
                libro = nombre_ant if not entrada else entrada.upper()

                texto = f"{funciones.AMARILLO}Introducir el nuevo nombre del autor: \033[s"
                funciones.escritura_lenta_print(texto)
                
                entrada = input("\033[u").strip()
                while entrada != "" and not re.match(patron_texto, entrada):
                    print(f"\n{funciones.ROJO}Autor inválido. Solo puede contener letras y espacios.{funciones.RESET}")
                    funciones.espereTecla()
                    print("\033[u\033[J", end="", flush=True)
                    entrada = input("\033[u").strip()
                    
                autor = autor_ant if not entrada else entrada.upper()

                texto = f"{funciones.AMARILLO}Introducir el nuevo nombre de la editorial: \033[s"
                funciones.escritura_lenta_print(texto)
                
                entrada = input("\033[u").strip()
                while entrada != "" and not re.match(patron_texto, entrada):
                    print(f"\n{funciones.ROJO}Editorial inválida. Solo debe contener letras y espacios.{funciones.RESET}")
                    funciones.espereTecla()
                    print("\033[u\033[J", end="", flush=True)
                    entrada = input("\033[u").strip()
                    
                editorial = editorial_ant if not entrada else entrada.upper()
                texto = f"{funciones.AMARILLO}Introducir el nuevo idioma del libro: \033[s"
                funciones.escritura_lenta_print(texto)
                
                entrada = input("\033[u").strip()
                while entrada != "" and not re.match(patron_texto, entrada):
                    print(f"\n{funciones.ROJO}Idioma inválido. Solo debe contener letras.{funciones.RESET}")
                    funciones.espereTecla()
                    print("\033[u\033[J", end="", flush=True)
                    entrada = input("\033[u").strip()
                    
                idioma = idioma_ant if not entrada else entrada.upper()

                respuesta = crud.actualizar(codigo, libro, autor, editorial, idioma, conexionBD)
            
                if respuesta:
                    funciones.accionExitosa()
                    funciones.menuLibros()
                else:
                    funciones.accionNoExitosa()
                    funciones.menuLibros()    
        elif opc=="no":
                funciones.menuLibros()
    else:        
            input(f"\n{funciones.ROJO}...¡No se encontró el libro con el código: {codigo}! Presione Enter...{funciones.RESET}")
            funciones.menuLibros()

def buscarLibrosAlumno(matricula, conexionBD):
    print("\033c")
    texto = f"{funciones.AZUL}\n" + "....:::: BUSCAR LIBROS ::::....".center(80) + f"\n{funciones.RESET}"
    funciones.escritura_lenta_print(texto)

    patron_nombre = r'^[A-ZÁÉÍÓÚÑa-záéíóúñ\s:.,\-#¿?¡!&"\'()$%+]{2,100}$'

    texto = f"{funciones.AMARILLO}Escribir el nombre del libro a buscar: {funciones.RESET}\033[s"
    funciones.escritura_lenta_print(texto)
    
    criterio = input("\033[u").strip()

    while criterio == "" or not re.match(patron_nombre, criterio):
        if criterio == "":
            print(f"\n{funciones.ROJO}El nombre del libro no puede estar vacío.{funciones.RESET}")
        else:
            print(f"\n{funciones.ROJO}Entrada inválida. Ingrese únicamente el nombre del libro.{funciones.RESET}")
            
        funciones.espereTecla()
        print("\033[u\033[J", end="", flush=True)
        criterio = input("\033[u").strip()

    criterio = criterio.upper()

    libros = crud.buscar(criterio, conexionBD)
    
    if len(libros) > 0:
        print(f"\n\t{funciones.AMARILLO}{'Código':<8}\t{'Título':<30}\t{'Autor':<25}\t{'Editorial':<20}\t{'Idioma':<12}{funciones.RESET}\n")
        for i in libros:
            print(f"\t{str(i[0]):<8}\t{str(i[1]):<30}\t{str(i[2]):<25}\t{str(i[3]):<20}\t{str(i[4]):<12}")
        
        funciones.espereTecla()
    else:        
        texto = f"\n{funciones.ROJO}...¡No se encontró ningún libro con el nombre '{criterio}'!...{funciones.RESET}"
        funciones.escritura_lenta_print(texto)
        funciones.espereTecla()

def generarReporteExcel(conexionBD):
    print("\033c")

    texto = (
        f"{funciones.AZUL}"
        + "....:::: GENERAR REPORTE DE LIBROS EN EXCEL ::::....".center(80)
        + f"\n{funciones.RESET}"
    )
    funciones.escritura_lenta_print(texto)

    try:
        libros = crud.consultar(conexionBD)

        if len(libros) == 0:
            print(
                f"\n{funciones.ROJO}"
                "No hay libros registrados para generar el reporte."
                f"{funciones.RESET}"
            )
            funciones.espereTecla()
            funciones.menuLibros()
            return

        libro_excel = Workbook()
        hoja = libro_excel.active
        hoja.title = "Reporte de libros"

        hoja.merge_cells("A1:E1")
        hoja["A1"] = "SISTEMA DE BIBLIOTECA UTD"
        hoja["A1"].font = Font(bold=True, size=16)
        hoja["A1"].alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

        hoja.merge_cells("A2:E2")
        hoja["A2"] = "REPORTE GENERAL DE LIBROS"
        hoja["A2"].font = Font(bold=True, size=14)
        hoja["A2"].alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

        fecha_actual = datetime.now().strftime("%d/%m/%Y %H:%M:%S")

        hoja.merge_cells("A3:E3")
        hoja["A3"] = f"Fecha de generación: {fecha_actual}"
        hoja["A3"].alignment = Alignment(horizontal="center")

        encabezados = [
            "Código",
            "Título",
            "Autor",
            "Editorial",
            "Idioma"
        ]

        fila_encabezados = 5

        borde_delgado = Side(
            style="thin",
            color="000000"
        )

        for columna, encabezado in enumerate(encabezados, start=1):
            celda = hoja.cell(
                row=fila_encabezados,
                column=columna,
                value=encabezado
            )

            celda.font = Font(
                bold=True,
                color="FFFFFF"
            )

            celda.fill = __import__(
                "openpyxl"
            ).styles.PatternFill(
                fill_type="solid",
                fgColor="1F4E78"
            )

            celda.alignment = Alignment(
                horizontal="center",
                vertical="center"
            )

            celda.border = Border(
                left=borde_delgado,
                right=borde_delgado,
                top=borde_delgado,
                bottom=borde_delgado
            )

        fila = 6
        contador_libros = 0

        for libro in libros:
            contador_libros += 1

            datos = [
                libro[0],
                libro[1],
                libro[2],
                libro[3],
                libro[4]
            ]

            for columna, dato in enumerate(datos, start=1):
                celda = hoja.cell(
                    row=fila,
                    column=columna,
                    value=dato
                )

                celda.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True
                )

                celda.border = Border(
                    left=borde_delgado,
                    right=borde_delgado,
                    top=borde_delgado,
                    bottom=borde_delgado
                )

            fila += 1

        fila_total = fila + 1

        hoja.cell(
            row=fila_total,
            column=1,
            value="Total de libros:"
        ).font = Font(bold=True)

        hoja.cell(
            row=fila_total,
            column=2,
            value=contador_libros
        ).font = Font(bold=True)

        anchos_columnas = {
            "A": 12,
            "B": 40,
            "C": 30,
            "D": 25,
            "E": 18
        }

        for columna, ancho in anchos_columnas.items():
            hoja.column_dimensions[columna].width = ancho

        hoja.row_dimensions[1].height = 25
        hoja.row_dimensions[2].height = 23
        hoja.row_dimensions[5].height = 25

        hoja.freeze_panes = "A6"
        hoja.auto_filter.ref = f"A5:E{fila - 1}"

        descargas = Path.home()/"Downloads"
        carpeta_reportes = descargas / "Reportes Libros"
        if not os.path.exists(carpeta_reportes):
            os.makedirs(carpeta_reportes)

        fecha_archivo = datetime.now().strftime("%Y%m%d_%H%M%S")

        nombre_archivo = (
            f"Reporte_Libros_{fecha_archivo}.xlsx"
        )

        ruta_archivo = os.path.join(
            carpeta_reportes,
            nombre_archivo
        )

        libro_excel.save(ruta_archivo)

        print(
            f"\n{funciones.VERDE}"
            "Reporte generado correctamente."
            f"{funciones.RESET}"
        )

        print(
            f"\n{funciones.AMARILLO}"
            f"Total de libros exportados: {contador_libros}"
            f"{funciones.RESET}"
        )

        print(
            f"\n{funciones.AMARILLO}"
            f"Archivo guardado en: {os.path.abspath(ruta_archivo)}"
            f"{funciones.RESET}"
        )

    except PermissionError:
        print(
            f"\n{funciones.ROJO}"
            "No se pudo guardar el reporte porque el archivo está abierto."
            "\nCierra el archivo de Excel e inténtalo nuevamente."
            f"{funciones.RESET}"
        )

    except Exception as error:
        print(
            f"\n{funciones.ROJO}"
            f"No se pudo generar el reporte: {error}"
            f"{funciones.RESET}"
        )

    funciones.espereTecla()
    funciones.menuLibros()