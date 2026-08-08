import re
import Usuarios.crud
import funciones 
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from pathlib import Path

def agregarUsuarios(conexionBD):
    print("\033c")
    texto = f"\n\t\t{funciones.AZUL}...:::: AGREGAR USUARIOS ::::...\n"
    funciones.escritura_lenta_print(texto)

    patron_matricula = r'^\d{1,11}$' 
    patron_nombre = r'^[A-ZÁÉÍÓÚÑa-záéíóúñ\s]{3,50}$'
    patron_password = r'^\S{4,20}$'

    texto = f"{funciones.AMARILLO}Introducir la matricula del usuario (solo números): {funciones.RESET}\033[s"
    funciones.escritura_lenta_print(texto)
    matricula = input("\033[u").strip()
    
    resp = Usuarios.crud.existe_matricula(matricula, conexionBD)
    
    while matricula == "" or not re.match(patron_matricula, matricula) or resp:
        if matricula == "":
            print(f"\n{funciones.ROJO}La matrícula no puede estar vacía.{funciones.RESET}")
        elif not re.match(patron_matricula, matricula):
            print(f"\n{funciones.ROJO}Formato inválido. Debe contener solo números con una longitud de 11 digitos.{funciones.RESET}")
        elif resp:
            print(f"\n{funciones.ROJO}¡La matrícula: {matricula} ya existe! Intente con otra.{funciones.RESET}")
            
        funciones.espereTecla()
        print("\033[u\033[J", end="", flush=True)
        matricula = input().strip()

        resp = Usuarios.crud.existe_matricula(matricula, conexionBD)

    texto = f"{funciones.AMARILLO}Introducir el nombre del usuario: {funciones.RESET}\033[s"
    funciones.escritura_lenta_print(texto)
    name = input("\033[u").strip()
    
    while name == "" or not re.match(patron_nombre, name):
        if name == "":
            print(f"\n{funciones.ROJO}El nombre no puede estar vacío.{funciones.RESET}")
        else:
            print(f"\n{funciones.ROJO}El nombre solo debe contener letras y espacios con una longitud maxima de 50 caracteres.{funciones.RESET}")
            
        funciones.espereTecla()
        print("\033[u\033[J", end="", flush=True)
        name = input().strip()
        
    name = name.upper()

    texto = f"{funciones.AMARILLO}Introducir la contraseña: {funciones.RESET}\033[s"
    funciones.escritura_lenta_print(texto)
    password = input("\033[u").strip()
    
    while password == "" or not re.match(patron_password, password):
        if password == "":
            print(f"\n{funciones.ROJO}La contraseña no puede estar vacía.{funciones.RESET}")
        else:
            print(f"\n{funciones.ROJO}La contraseña no debe contener espacios y debe tener una longitud minima de 4 caracteres y maxima de 20.{funciones.RESET}")
            
        funciones.espereTecla()
        print("\033[u\033[J", end="", flush=True)
        password = input().strip()

    roles_permitidos = ["TRABAJADOR", "ALUMNO"]
    
    texto = f"{funciones.AMARILLO}Introducir el rol (TRABAJADOR/ALUMNO): {funciones.RESET}\033[s"
    funciones.escritura_lenta_print(texto)
    rol = input("\033[u").upper().strip()

    while rol not in roles_permitidos:
        print(f"\n{funciones.ROJO}Rol inválido. Opciones válidas: TRABAJADOR O ALUMNO.{funciones.RESET}")
        funciones.espereTecla()
        print("\033[u\033[J", end="", flush=True)
        rol = input().upper().strip()

    respuesta = Usuarios.crud.insertar(matricula, name, password, rol, conexionBD)
    if respuesta: 
        funciones.accionExitosa()
        funciones.menuUsuarios()
    else:
        funciones.accionNoExitosa()
        funciones.menuUsuarios()

def mostrarUsuarios(conexionBD): 
    print("\033c")
    texto = f"{funciones.AZUL}\n\t\t\t\t...:::: MOSTRAR USUARIOS ::::...\n{funciones.AMARILLO}"
    funciones.escritura_lenta_print(texto)
    
    usuarios = Usuarios.crud.consultar(conexionBD)
    
    if len(usuarios) > 0:
        print(f"\t{funciones.AMARILLO}{'Matrícula':<15}{'Nombre':<25}{'Contraseña':<20}{'Rol':<15}{funciones.RESET}\n")
        
        for i in usuarios:
            mat = str(i[0])[:15]
            nombre = str(i[1])[:24]
            pwd = str(i[2])[:19]
            rol = str(i[3])[:14]
            print(f"\t{mat:<15}{nombre:<25}{pwd:<20}{rol:<15}")
            
        print("\n")
        funciones.espereTecla()
        funciones.menuUsuarios()
    else:        
        texto = f"\n{funciones.ROJO}...¡No hay usuarios que mostrar!...{funciones.RESET}\n"
        funciones.escritura_lenta_print(texto)
        funciones.espereTecla()
        funciones.menuUsuarios()

def buscarUsuarios(conexionBD):
    print("\033c")
    texto = f"{funciones.AZUL}\n\t\t\t\t...:::: BUSCAR USUARIOS ::::...\n"
    funciones.escritura_lenta_print(texto)
    
    texto = f"{funciones.AMARILLO}Escribir la matrícula o nombre del usuario: {funciones.RESET}\033[s"
    funciones.escritura_lenta_print(texto)
    criterio = input("\033[u").strip()

    while criterio == "":
        print(f"\n{funciones.ROJO}El nombre o matricula de búsqueda no puede estar vacío.{funciones.RESET}")
        funciones.espereTecla()
        print("\033[u\033[J", end="", flush=True)
        criterio = input().strip()

    criterio = criterio.upper()
    usuarios = Usuarios.crud.buscar(criterio, conexionBD)
    
    if len(usuarios) > 0:
        cont_usuarios=0
        for u in usuarios:
            cont_usuarios+=1
        print(f"\t{funciones.VERDE}Se encontraron {cont_usuarios} registros{funciones.RESET}\n")
        print(f"\n\t{funciones.AMARILLO}{'Matrícula':<15}{'Nombre':<25}{'Contraseña':<20}{'Rol':<15}{funciones.RESET}\n")

        for u in usuarios:
            mat = str(u[0])[:15]
            nombre = str(u[1])[:24]
            pwd = str(u[2])[:19]
            rol = str(u[3])[:14]
            print(f"\t{mat:<15}{nombre:<25}{pwd:<20}{rol:<15}")
            
        print("\n")
        funciones.espereTecla()
        funciones.menuUsuarios()
    else:        
        texto = f"\n{funciones.ROJO}...¡No se encontró ningún usuario con esa matrícula o nombre!...{funciones.RESET}\n"
        funciones.escritura_lenta_print(texto)
        funciones.espereTecla()
        funciones.menuUsuarios()

def borrarUsuarios(conexionBD):
    print("\033c")
    texto = f"{funciones.AZUL}\n\t\t\t\t...:::: BORRAR USUARIO ::::...\n{funciones.RESET}"
    funciones.escritura_lenta_print(texto)

    patron_matricula = r'^\d{1,11}$' 

    texto = f"{funciones.AMARILLO}Escribir la matrícula del usuario a borrar: {funciones.RESET}\033[s"
    funciones.escritura_lenta_print(texto)
    
    matricula = input("\033[u").strip()

    while matricula == "" or not re.match(patron_matricula, matricula):
        print(f"\n{funciones.ROJO}Entrada inválida. Debe ingresar únicamente los números de la matrícula.{funciones.RESET}")
        funciones.espereTecla()
        print("\033[u\033[J", end="", flush=True)
        matricula = input().strip()

    usuarios = Usuarios.crud.buscar(matricula, conexionBD)
    
    if len(usuarios) > 0:
        cursor = conexionBD.cursor()
        cursor.execute("SELECT COUNT(*) FROM prestamos WHERE id_usuario = %s", (matricula,))
        prestamos_asociados = cursor.fetchone()[0]

        if prestamos_asociados > 0:
            input(f"\n{funciones.ROJO}...¡No se puede borrar! El usuario tiene préstamos registrados en el sistema. Presione Enter...{funciones.RESET}")
            funciones.menuUsuarios()
            return

        print(f"\n\t{funciones.AMARILLO}{'Matrícula':<15}{'Nombre':<25}{'Contraseña':<20}{'Rol':<15}{funciones.RESET}\n")
        for u in usuarios:
            mat = str(u[0])[:15]
            nombre = str(u[1])[:24]
            pwd = str(u[2])[:19]
            rol = str(u[3])[:14]
            print(f"\t{mat:<15}{nombre:<25}{pwd:<20}{rol:<15}")
        
        texto = f"\n{funciones.ROJO}¿Deseas borrar al usuario mostrado (Si/No)? {funciones.RESET}\033[s"
        funciones.escritura_lenta_print(texto)
        opc = input("\033[u").lower().strip()

        while opc not in ["si", "no"]:
            print(f"\n{funciones.ROJO}Opción inválida. Escriba unicamente SI o NO.{funciones.RESET}")
            funciones.espereTecla()
            print("\033[u\033[J", end="", flush=True)
            opc = input().lower().strip()
                
        if opc == "si":       
            respuesta = Usuarios.crud.borrar(matricula, conexionBD)
            
            if respuesta:
                funciones.accionExitosa()
            else:
                funciones.accionNoExitosa()
                
            funciones.menuUsuarios() 
        else:
            funciones.menuUsuarios()
    else:        
        input(f"\n{funciones.ROJO}...¡No se encontró ningún usuario registrado con la matrícula: {matricula}! Presione Enter...{funciones.RESET}")
        funciones.menuUsuarios()

def limpiarUsuarios(conexionBD):
    print("\033c")
    texto = f"{funciones.AZUL}\n\t\t\t\t...:::: LIMPIAR USUARIOS ::::...\n{funciones.RESET}"
    funciones.escritura_lenta_print(texto)
    
    usuarios = Usuarios.crud.consultar(conexionBD)
    
    if len(usuarios) > 0:
        cursor = conexionBD.cursor()
        cursor.execute("SELECT COUNT(*) FROM prestamos")
        total_prestamos = cursor.fetchone()[0]

        if total_prestamos > 0:
            input(f"\n{funciones.ROJO}...¡No se puede vaciar! Hay préstamos registrados. Presione Enter...{funciones.RESET}")
            funciones.menuUsuarios()
            return
        texto = f"{funciones.ROJO}¿Deseas borrar TODOS los usuarios (Si/No)? {funciones.RESET}\033[s"
        funciones.escritura_lenta_print(texto)
        opc = input("\033[u").lower().strip()

        while opc not in ["si", "no"]:
            print(f"\n{funciones.ROJO}Opción inválida. Escriba unicamente SI o NO.{funciones.RESET}")
            funciones.espereTecla()
            print("\033[u\033[J", end="", flush=True)
            opc = input().lower().strip()
                
        if opc == "si":       
            respuesta = Usuarios.crud.vaciar(conexionBD)
            
            if respuesta:
                funciones.accionExitosa()
            else:
                funciones.accionNoExitosa()
                
            funciones.menuUsuarios()       
        else:
            funciones.menuUsuarios()
    else:        
        input(f"\n{funciones.ROJO}...¡No hay usuarios que borrar! Presione Enter...{funciones.RESET}") 
        funciones.menuUsuarios()

def modificarUsuariosAlumno(matricula,conexionBD):
    print("\033c")
    texto = f"{funciones.AZUL}\n\t\t\t\t...:::: MODIFICAR MIS DATOS ::::...\n"
    funciones.escritura_lenta_print(texto)

    usuarios = Usuarios.crud.buscar(matricula, conexionBD)
    usuario_actual = usuarios[0] 
    nombre_ant = usuario_actual[1]
    password_ant = usuario_actual[2]
    rol = usuario_actual[3]
    
    texto = f"{funciones.ROJO}¿Deseas actualizar tu información (Si/No)? {funciones.RESET}\033[s"
    funciones.escritura_lenta_print(texto)
    opc = input("\033[u").lower().strip()
        
    while opc not in ["si", "no"]:
            print(f"\n{funciones.ROJO}Opción inválida. Escriba unicamente SI o NO.{funciones.RESET}")
            funciones.espereTecla()
            print("\033[u\033[J", end="", flush=True)
            opc = input().lower().strip()
            
    if opc == "si":
            print(f"\n{funciones.AZUL}(Presiona ENTER sin escribir nada para conservar el valor actual){funciones.RESET}\n")

            patron_nombre = r'^[A-ZÁÉÍÓÚÑa-záéíóúñ\s]{3,50}$'
            patron_password = r'^\S{4,20}$'

            texto = f"{funciones.AMARILLO}Introducir el nuevo nombre: {funciones.RESET}\033[s"
            funciones.escritura_lenta_print(texto)
            entrada = input("\033[u").strip()
            
            while entrada != "" and not re.match(patron_nombre, entrada):
                print(f"\n{funciones.ROJO}Nombre inválido. Solo puede contener letras y espacios con una longitud maxima de 50 caracteres.{funciones.RESET}")
                funciones.espereTecla()
                print("\033[u\033[J", end="", flush=True)
                entrada = input().strip()
                
            name = nombre_ant if not entrada else entrada.upper()

            texto = f"{funciones.AMARILLO}Introducir la nueva contraseña: {funciones.RESET}\033[s"
            funciones.escritura_lenta_print(texto)
            entrada = input("\033[u").strip()
            
            while entrada != "" and not re.match(patron_password, entrada):
                print(f"\n{funciones.ROJO}Contraseña inválida.No debe de contener espacios y la longitud minima es de 4 caracteres y maxima de 20.{funciones.RESET}")
                funciones.espereTecla()
                print("\033[u\033[J", end="", flush=True)
                entrada = input().strip()
                
            password = password_ant if not entrada else entrada

            respuesta = Usuarios.crud.actualizar(matricula, name, password, rol, conexionBD)
        
            if respuesta:
                funciones.accionExitosa()
                funciones.menuPrincipalAlumno(matricula)
            else:
                funciones.accionNoExitosa()
                funciones.menuPrincipalAlumno(matricula)    
    else:
            funciones.menuPrincipalAlumno(matricula)

def modificarUsuarios(conexionBD):
    print("\033c")
    texto = f"{funciones.AZUL}\n\t\t\t\t...:::: MODIFICAR USUARIOS ::::...\n"
    funciones.escritura_lenta_print(texto)
    
    patron_matricula = r'^\d{1,11}$'

    texto = f"{funciones.AMARILLO}Escribir la matrícula del usuario que quieres actualizar: {funciones.RESET}\033[s"
    funciones.escritura_lenta_print(texto)
    
    matricula = input("\033[u").strip()
    
    while matricula == "" or not re.match(patron_matricula, matricula):
        print(f"\n{funciones.ROJO}Entrada inválida. Ingrese solo números con una longitud maxima de 11 digitos.{funciones.RESET}")
        funciones.espereTecla()
        print("\033[u\033[J", end="", flush=True)
        matricula = input().strip()
    
    usuarios = Usuarios.crud.buscar(matricula, conexionBD)
    
    if len(usuarios) > 0:
        usuario_actual = usuarios[0] 
        nombre_ant = usuario_actual[1]
        password_ant = usuario_actual[2]
        rol_ant = usuario_actual[3]
        print(f"\n\t{funciones.AMARILLO}{'Matrícula':<15}{'Nombre':<25}{'Contraseña':<20}{'Rol':<15}{funciones.RESET}\n")
        mat = str(usuario_actual[0])[:15]
        nom = str(nombre_ant)[:24]
        pwd = str(password_ant)[:19]
        rl = str(rol_ant)[:14]
        print(f"\t{mat:<15}{nom:<25}{pwd:<20}{rl:<15}\n")
        
        texto = f"{funciones.ROJO}¿Deseas actualizar a este usuario (Si/No)? {funciones.RESET}\033[s"
        funciones.escritura_lenta_print(texto)
        opc = input("\033[u").lower().strip()
        
        while opc not in ["si", "no"]:
            print(f"\n{funciones.ROJO}Opción inválida. Escriba unicamente SI o NO.{funciones.RESET}")
            funciones.espereTecla()
            print("\033[u\033[J", end="", flush=True)
            opc = input().lower().strip()
            
        if opc == "si":
            print(f"\n{funciones.AZUL}(Presiona ENTER sin escribir nada para conservar el valor actual){funciones.RESET}\n")

            patron_nombre = r'^[A-ZÁÉÍÓÚÑa-záéíóúñ\s]{3,50}$'
            patron_password = r'^\S{4,20}$'
            roles_permitidos = ["TRABAJADOR", "ALUMNO"]

            texto = f"{funciones.AMARILLO}Introducir el nuevo nombre: {funciones.RESET}\033[s"
            funciones.escritura_lenta_print(texto)
            entrada = input("\033[u").strip()
            
            while entrada != "" and not re.match(patron_nombre, entrada):
                print(f"\n{funciones.ROJO}Nombre inválido. Solo letras y espacios con una longitud minima de 3 caracteres y una maxima de 50.{funciones.RESET}")
                funciones.espereTecla()
                print("\033[u\033[J", end="", flush=True)
                entrada = input().strip()
                
            name = nombre_ant if not entrada else entrada.upper()

            texto = f"{funciones.AMARILLO}Introducir la nueva contraseña: {funciones.RESET}\033[s"
            funciones.escritura_lenta_print(texto)
            entrada = input("\033[u").strip()
            
            while entrada != "" and not re.match(patron_password, entrada):
                print(f"\n{funciones.ROJO}Contraseña inválida.No debe de contener espacios y debe de tener una longitud minima de 4 caracteres y maxima de 20{funciones.RESET}")
                funciones.espereTecla()
                print("\033[u\033[J", end="", flush=True)
                entrada = input().strip()
                
            password = password_ant if not entrada else entrada

            texto = f"{funciones.AMARILLO}Introducir nuevo rol (TRABAJADOR/ALUMNO): {funciones.RESET}\033[s"
            funciones.escritura_lenta_print(texto)
            entrada = input("\033[u").upper().strip()
            
            while entrada != "" and entrada not in roles_permitidos:
                print(f"\n{funciones.ROJO}Rol inválido. Opciones: TRABAJADOR O ALUMNO.{funciones.RESET}")
                funciones.espereTecla()
                print("\033[u\033[J", end="", flush=True)
                entrada = input().upper().strip()
                
            rol = rol_ant if not entrada else entrada
            respuesta = Usuarios.crud.actualizar(matricula, name, password, rol, conexionBD)
        
            if respuesta:
                funciones.accionExitosa()
                funciones.menuUsuarios()
            else:
                funciones.accionNoExitosa()
                funciones.menuUsuarios()    
        else:
            funciones.menuUsuarios()
    else:        
        input(f"\n{funciones.ROJO}...¡No se encontró ningún usuario con la matrícula: {matricula}! Presione Enter...{funciones.RESET}")
        funciones.menuUsuarios()

def generarReporteExcel(conexionBD):
    print("\033c")

    texto = (
        f"{funciones.AZUL}"
        + "....:::: GENERAR REPORTE DE USUARIOS EN EXCEL ::::....".center(80)
        + f"\n{funciones.RESET}"
    )
    funciones.escritura_lenta_print(texto)

    try:
        lista_usuarios = Usuarios.crud.consultar(conexionBD)

        if len(lista_usuarios) == 0:
            print(
                f"\n{funciones.ROJO}"
                "No hay usuarios registrados para generar el reporte."
                f"{funciones.RESET}"
            )
            funciones.espereTecla()
            funciones.menuUsuarios()
            return

        archivo_excel = Workbook()
        hoja = archivo_excel.active
        hoja.title = "Reporte de usuarios"

        hoja.merge_cells("A1:C1")
        hoja["A1"] = "SISTEMA DE BIBLIOTECA UTD"
        hoja["A1"].font = Font(bold=True, size=16)
        hoja["A1"].alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

        hoja.merge_cells("A2:C2")
        hoja["A2"] = "REPORTE GENERAL DE USUARIOS"
        hoja["A2"].font = Font(bold=True, size=14)
        hoja["A2"].alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

        fecha_actual = datetime.now().strftime("%d/%m/%Y %H:%M:%S")

        hoja.merge_cells("A3:C3")
        hoja["A3"] = f"Fecha de generación: {fecha_actual}"
        hoja["A3"].alignment = Alignment(horizontal="center")

        encabezados = [
            "Matrícula",
            "Nombre",
            "Rol"
        ]

        relleno_encabezado = PatternFill(
            fill_type="solid",
            fgColor="1F4E78"
        )

        borde_delgado = Side(
            style="thin",
            color="000000"
        )

        borde = Border(
            left=borde_delgado,
            right=borde_delgado,
            top=borde_delgado,
            bottom=borde_delgado
        )

        fila_encabezados = 5

        for columna, encabezado in enumerate(encabezados, start=1):
            celda = hoja.cell(
                row=fila_encabezados,
                column=columna,
                value=encabezado
            )

            celda.font = Font(
                bold=True,
                color="FFFFFF"
            )

            celda.fill = relleno_encabezado

            celda.alignment = Alignment(
                horizontal="center",
                vertical="center"
            )

            celda.border = borde

        fila = 6
        contador_usuarios = 0

        for usuario in lista_usuarios:
            contador_usuarios += 1

            datos_usuario = {
                "matricula": str(usuario[0]),
                "nombre": usuario[1],
                "rol": usuario[3]
            }

            for columna, dato in enumerate(datos_usuario.values(), start=1):
                celda = hoja.cell(
                    row=fila,
                    column=columna,
                    value=dato
                )

                celda.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True
                )

                celda.border = borde

            fila += 1

        fila_total = fila + 1

        hoja.cell(
            row=fila_total,
            column=1,
            value="Total de usuarios:"
        ).font = Font(bold=True)

        hoja.cell(
            row=fila_total,
            column=2,
            value=contador_usuarios
        ).font = Font(bold=True)

        hoja.column_dimensions["A"].width = 18
        hoja.column_dimensions["B"].width = 40
        hoja.column_dimensions["C"].width = 20

        hoja.row_dimensions[1].height = 25
        hoja.row_dimensions[2].height = 23
        hoja.row_dimensions[5].height = 25

        hoja.freeze_panes = "A6"
        hoja.auto_filter.ref = f"A5:C{fila - 1}"

        descargas = Path.home()/"Downloads"
        carpeta_reportes_user = descargas / "Reportes usuarios"

        if not os.path.exists(carpeta_reportes_user):
            os.makedirs(carpeta_reportes_user)

        fecha_archivo = datetime.now().strftime("%Y%m%d_%H%M%S")

        nombre_archivo = (
            f"Reporte_Usuarios_{fecha_archivo}.xlsx"
        )

        ruta_archivo = os.path.join(
            carpeta_reportes_user,
            nombre_archivo
        )

        archivo_excel.save(ruta_archivo)

        print(
            f"\n{funciones.VERDE}"
            "Reporte de usuarios generado correctamente."
            f"{funciones.RESET}"
        )

        print(
            f"\n{funciones.AMARILLO}"
            f"Total de usuarios exportados: {contador_usuarios}"
            f"{funciones.RESET}"
        )

        print(
            f"\n{funciones.AMARILLO}"
            f"Archivo guardado en:\n{os.path.abspath(ruta_archivo)}"
            f"{funciones.RESET}"
        )

    except PermissionError:
        print(
            f"\n{funciones.ROJO}"
            "No se pudo guardar el reporte porque el archivo está abierto."
            "\nCierra el archivo de Excel e inténtalo nuevamente."
            f"{funciones.RESET}"
        )

    except Exception as error:
        print(
            f"\n{funciones.ROJO}"
            f"No se pudo generar el reporte: {error}"
            f"{funciones.RESET}"
        )

    funciones.espereTecla()
    funciones.menuUsuarios()
    